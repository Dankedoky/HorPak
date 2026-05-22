from fastapi import FastAPI, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
import os
from dotenv import load_dotenv
import re
import io
from datetime import datetime, date, time, timedelta
import models, schemas, database
from auth import create_token, LoginRequest, get_current_user
from database import engine, get_db
from linebot.v3 import WebhookParser
from linebot.v3.messaging import Configuration, AsyncApiClient, AsyncMessagingApi, ReplyMessageRequest, TextMessage, PushMessageRequest, AsyncMessagingApiBlob, FlexMessage, FlexContainer
from linebot.v3.exceptions import InvalidSignatureError
from linebot.v3.webhooks import MessageEvent, TextMessageContent

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"))

models.Base.metadata.create_all(bind=engine)

# LINE Config
channel_secret = os.getenv("LINE_CHANNEL_SECRET")
channel_access_token = os.getenv("LINE_CHANNEL_ACCESS_TOKEN")
parser = WebhookParser(channel_secret) if channel_secret else None

_line_bot_api = None
_line_blob_api = None

def get_line_bot_api() -> AsyncMessagingApi:
    global _line_bot_api
    if _line_bot_api is None:
        configuration = Configuration(access_token=channel_access_token)
        async_api_client = AsyncApiClient(configuration)
        _line_bot_api = AsyncMessagingApi(async_api_client)
    return _line_bot_api

def get_line_blob_api() -> AsyncMessagingApiBlob:
    global _line_blob_api
    if _line_blob_api is None:
        configuration = Configuration(access_token=channel_access_token)
        async_api_client = AsyncApiClient(configuration)
        _line_blob_api = AsyncMessagingApiBlob(async_api_client)
    return _line_blob_api


from fastapi.middleware.cors import CORSMiddleware

app = FastAPI(title="Sovereign Accounting API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

def get_current_billing_month(ref_date_str: str = None) -> str:
    from datetime import datetime, timedelta
    if ref_date_str:
        try:
            date_part = ref_date_str.split(" ")[0].split("T")[0]
            if "/" in date_part:
                parts = date_part.split("/")
                ref_date = datetime(int(parts[2]), int(parts[1]), int(parts[0]))
            else:
                ref_date = datetime.strptime(date_part, "%Y-%m-%d")
        except Exception:
            ref_date = datetime.now()
    else:
        ref_date = datetime.now()
        
    if ref_date.day >= 25:
        return ref_date.strftime("%Y-%m")
    else:
        first_of_this_month = ref_date.replace(day=1)
        prev_month_last_day = first_of_this_month - timedelta(days=1)
        return prev_month_last_day.strftime("%Y-%m")


def create_dorm_bill_flex(room, current_month_key: str) -> dict:
    import json
    
    rate = room.rate or 0.0
    water_cost = room.water_cost or 0.0
    electric_cost = room.electric_cost or 0.0
    cleaning_fee = room.cleaning_fee or 0.0
    other_fee = room.other_fee or 0.0
    fine_cost = room.fine_cost or 0.0
    
    water_meter_prev = room.water_meter_prev or 0.0
    water_meter = room.water_meter or 0.0
    electricity_meter_prev = room.electricity_meter_prev or 0.0
    electricity_meter = room.electricity_meter or 0.0
    late_days = room.late_days or 0
    
    total_amount = rate + water_cost + electric_cost + cleaning_fee + other_fee + fine_cost
    
    parts = current_month_key.split("-")
    display_month = f"{parts[1]}/{parts[0]}" if len(parts) == 2 else current_month_key
    
    body_contents = [
        # Metadata section
        {
            "type": "box",
            "layout": "vertical",
            "spacing": "sm",
            "contents": [
                {
                    "type": "box",
                    "layout": "horizontal",
                    "contents": [
                        {"type": "text", "text": "🏢 ห้องพัก:", "color": "#718096", "size": "sm", "flex": 3},
                        {"type": "text", "text": f"ห้อง {room.number}", "weight": "bold", "color": "#2D3748", "size": "sm", "align": "end", "flex": 5}
                    ]
                },
                {
                    "type": "box",
                    "layout": "horizontal",
                    "contents": [
                        {"type": "text", "text": "👤 ผู้เช่า:", "color": "#718096", "size": "sm", "flex": 3},
                        {"type": "text", "text": f"คุณ {room.tenant or 'ผู้เช่า'}", "weight": "bold", "color": "#2D3748", "size": "sm", "align": "end", "flex": 5}
                    ]
                },
                {
                    "type": "box",
                    "layout": "horizontal",
                    "contents": [
                        {"type": "text", "text": "📅 ประจำรอบบิล:", "color": "#718096", "size": "sm", "flex": 4},
                        {"type": "text", "text": display_month, "weight": "bold", "color": "#2D3748", "size": "sm", "align": "end", "flex": 4}
                    ]
                }
            ]
        },
        {"type": "separator", "margin": "lg", "color": "#E2E8F0"},
        # Breakdown section
        {
            "type": "box",
            "layout": "vertical",
            "margin": "lg",
            "spacing": "md",
            "contents": [
                # Rent Room
                {
                    "type": "box",
                    "layout": "horizontal",
                    "contents": [
                        {"type": "text", "text": "💵 ค่าเช่าห้อง:", "color": "#4A5568", "size": "sm", "flex": 5},
                        {"type": "text", "text": f"{rate:,.2f} บาท", "weight": "bold", "color": "#2D3748", "size": "sm", "align": "end", "flex": 3}
                    ]
                },
                # Water cost
                {
                    "type": "box",
                    "layout": "vertical",
                    "contents": [
                        {
                            "type": "box",
                            "layout": "horizontal",
                            "contents": [
                                {"type": "text", "text": "💧 ค่าน้ำประปา:", "color": "#4A5568", "size": "sm", "flex": 5},
                                {"type": "text", "text": f"{water_cost:,.2f} บาท", "weight": "bold", "color": "#2D3748", "size": "sm", "align": "end", "flex": 3}
                            ]
                        },
                        {
                            "type": "text",
                            "text": f"   (มิเตอร์ {water_meter_prev:g} -> {water_meter:g})",
                            "color": "#718096",
                            "size": "xs",
                            "margin": "xs"
                        }
                    ]
                },
                # Electric cost
                {
                    "type": "box",
                    "layout": "vertical",
                    "contents": [
                        {
                            "type": "box",
                            "layout": "horizontal",
                            "contents": [
                                {"type": "text", "text": "⚡️ ค่าไฟฟ้า:", "color": "#4A5568", "size": "sm", "flex": 5},
                                {"type": "text", "text": f"{electric_cost:,.2f} บาท", "weight": "bold", "color": "#2D3748", "size": "sm", "align": "end", "flex": 3}
                            ]
                        },
                        {
                            "type": "text",
                            "text": f"   (มิเตอร์ {electricity_meter_prev:g} -> {electricity_meter:g})",
                            "color": "#718096",
                            "size": "xs",
                            "margin": "xs"
                        }
                    ]
                }
            ]
        }
    ]
    
    # Append optional fields dynamically
    if cleaning_fee > 0:
        body_contents[2]["contents"].append({
            "type": "box",
            "layout": "horizontal",
            "contents": [
                {"type": "text", "text": "🧹 ค่าทำความสะอาด:", "color": "#4A5568", "size": "sm", "flex": 5},
                {"type": "text", "text": f"{cleaning_fee:,.2f} บาท", "weight": "bold", "color": "#2D3748", "size": "sm", "align": "end", "flex": 3}
            ]
        })
        
    if other_fee > 0:
        body_contents[2]["contents"].append({
            "type": "box",
            "layout": "horizontal",
            "contents": [
                {"type": "text", "text": "📦 ค่าบริการอื่นๆ:", "color": "#4A5568", "size": "sm", "flex": 5},
                {"type": "text", "text": f"{other_fee:,.2f} บาท", "weight": "bold", "color": "#2D3748", "size": "sm", "align": "end", "flex": 3}
            ]
        })
        
    if fine_cost > 0:
        body_contents[2]["contents"].append({
            "type": "box",
            "layout": "horizontal",
            "contents": [
                {"type": "text", "text": f"⚠️ ค่าปรับล่าช้า ({late_days} วัน):", "color": "#E53E3E", "size": "sm", "flex": 5},
                {"type": "text", "text": f"{fine_cost:,.2f} บาท", "weight": "bold", "color": "#E53E3E", "size": "sm", "align": "end", "flex": 3}
            ]
        })
        
    body_contents.append({"type": "separator", "margin": "lg", "color": "#E2E8F0"})
    body_contents.append({
        "type": "box",
        "layout": "horizontal",
        "margin": "lg",
        "contents": [
            {"type": "text", "text": "💰 ยอดรวมที่ต้องชำระ:", "weight": "bold", "color": "#1A365D", "size": "md", "flex": 5},
            {"type": "text", "text": f"{total_amount:,.2f} บาท", "weight": "bold", "color": "#E53E3E", "size": "md", "align": "end", "flex": 3}
        ]
    })
    
    flex_data = {
        "type": "bubble",
        "styles": {
            "header": {"backgroundColor": "#1A365D"}
        },
        "header": {
            "type": "box",
            "layout": "vertical",
            "contents": [
                {"type": "text", "text": "🧾 ใบแจ้งยอดค่าเช่าหอพัก 🧾", "weight": "bold", "color": "#FFFFFF", "size": "md", "align": "center"}
            ]
        },
        "body": {
            "type": "box",
            "layout": "vertical",
            "contents": body_contents
        },
        "footer": {
            "type": "box",
            "layout": "vertical",
            "spacing": "xs",
            "contents": [
                {
                    "type": "text",
                    "text": "📌 ชำระภายในวันที่ 5 ของเดือน หลังจากนี้มีค่าปรับวันละ 100 บาท",
                    "color": "#E53E3E",
                    "size": "xs",
                    "wrap": True,
                    "align": "center",
                    "weight": "bold"
                },
                {"type": "separator", "margin": "md", "color": "#E2E8F0"},
                {
                    "type": "text",
                    "text": "🏦 วิธีการชำระเงิน:",
                    "weight": "bold",
                    "color": "#1A365D",
                    "size": "sm",
                    "margin": "md"
                },
                {
                    "type": "text",
                    "text": "กรุณาโอนเงินเข้าบัญชีธนาคาร และส่งสลิปโอนเงินเข้ามาในแชท LINE OA นี้ เพื่อให้ระบบสแกนสลิปและปรับปรุงยอดโดยอัตโนมัติครับ 🙏",
                    "color": "#4A5568",
                    "size": "xs",
                    "wrap": True,
                    "margin": "xs"
                },
                {
                    "type": "button",
                    "style": "primary",
                    "color": "#1A365D",
                    "action": {
                        "type": "message",
                        "label": "📸 แจ้งโอนเงิน",
                        "text": "แจ้งโอนเงิน"
                    },
                    "margin": "lg"
                }
            ]
        }
    }
    return flex_data


def create_consolidated_bill_flex(tenant_name: str, room_number: str, unpaid_items: list, grand_total: float) -> dict:
    import json
    
    body_contents = [
        {
            "type": "text",
            "text": "รายการยอดค้างชำระสะสมในระบบของคุณมีดังนี้ค่ะ:",
            "color": "#4A5568",
            "size": "sm",
            "wrap": True
        },
        {"type": "separator", "margin": "md"}
    ]
    
    for idx, item in enumerate(unpaid_items, 1):
        item_box_contents = [
            {
                "type": "text",
                "text": f"[{idx}] {item['title']}",
                "weight": "bold",
                "color": "#2B6CB0",
                "size": "sm",
                "wrap": True
            }
        ]
        
        detail_box = {
            "type": "box",
            "layout": "vertical",
            "margin": "sm",
            "spacing": "xs",
            "contents": []
        }
        for d in item["details"]:
            detail_box["contents"].append({
                "type": "text",
                "text": d,
                "color": "#718096",
                "size": "xs",
                "margin": "xs"
            })
            
        item_box_contents.append(detail_box)
        item_box_contents.append({
            "type": "box",
            "layout": "horizontal",
            "margin": "xs",
            "contents": [
                {"type": "text", "text": "ยอดที่ต้องชำระ", "color": "#718096", "size": "xs"},
                {"type": "text", "text": f"{item['amount']:,.2f} บาท", "weight": "bold", "color": "#2D3748", "size": "sm", "align": "end"}
            ]
        })
        
        body_contents.append({
            "type": "box",
            "layout": "vertical",
            "margin": "md",
            "spacing": "xs",
            "contents": item_box_contents
        })
        
        if idx < len(unpaid_items):
            body_contents.append({"type": "separator", "margin": "md"})
            
    body_contents.append({"type": "separator", "margin": "lg"})
    body_contents.append({
        "type": "box",
        "layout": "horizontal",
        "margin": "lg",
        "contents": [
            {"type": "text", "text": "ยอดรวมสุทธิทั้งหมด", "weight": "bold", "color": "#2B6CB0", "size": "md"},
            {"type": "text", "text": f"{grand_total:,.2f} บาท", "weight": "bold", "color": "#E53E3E", "size": "lg", "align": "end"}
        ]
    })
    
    flex_data = {
        "type": "bubble",
        "styles": {
            "header": {"backgroundColor": "#2B6CB0"}
        },
        "header": {
            "type": "box",
            "layout": "vertical",
            "contents": [
                {"type": "text", "text": "สรุปยอดค้างชำระทั้งหมด", "weight": "bold", "color": "#FFFFFF", "size": "lg"},
                {"type": "text", "text": f"คุณ {tenant_name} | ห้อง {room_number}", "color": "#E2E8F0", "size": "sm", "margin": "xs"}
            ]
        },
        "body": {
            "type": "box",
            "layout": "vertical",
            "contents": body_contents
        },
        "footer": {
            "type": "box",
            "layout": "vertical",
            "spacing": "xs",
            "contents": [
                {
                    "type": "text",
                    "text": "💡 ท่านสามารถเลือกโอนชำระเงินยอดใดก่อนก็ได้ตามสะดวกนะคะ ระบบ AI จะสแกนตัดยอดจ่ายบิลใบนั้นให้ทันทีค่ะ",
                    "color": "#718096",
                    "size": "xs",
                    "wrap": True,
                    "align": "center"
                },
                {
                    "type": "button",
                    "style": "primary",
                    "color": "#2B6CB0",
                    "action": {
                        "type": "message",
                        "label": "📸 แจ้งโอนเงิน",
                        "text": "แจ้งโอนเงิน"
                    },
                    "margin": "md"
                }
            ]
        }
    }
    return flex_data


# ==========================================
# Auth Endpoint
# ==========================================
@app.post("/auth/login")
def login(req: LoginRequest):
    admin_password = os.getenv('ADMIN_PASSWORD')
    if not admin_password:
        raise HTTPException(status_code=500, detail="ADMIN_PASSWORD is not configured on the server")
    if req.password != admin_password:
        raise HTTPException(status_code=401, detail="รหัสผ่านไม่ถูกต้อง")
    token = create_token({"role": "admin"})
    return {"access_token": token, "token_type": "bearer"}

@app.on_event("startup")
def seed_business_units():
    # Schema check and self-healing
    db = next(get_db())
    schema_mismatch = False
    try:
        # Check if new columns exist by running a query
        db.query(models.DormRoom).first()
        db.query(models.RentalHouse).first()
        db.query(models.DormPayment).first()
        db.query(models.HousePayment).first()
    except Exception as e:
        err_str = str(e).lower()
        if "column" in err_str or "does not exist" in err_str or "no such column" in err_str or "relation" in err_str or "table" in err_str:
            schema_mismatch = True
            print(f"Database schema mismatch detected: {e}. Initiating self-healing...")
    finally:
        db.close()

    if schema_mismatch:
        try:
            # Drop and recreate tables to heal schema
            models.Base.metadata.drop_all(bind=engine)
            models.Base.metadata.create_all(bind=engine)
            print("Self-healing database schema completed successfully.")
        except Exception as err:
            print(f"Critical error during database self-healing: {err}")

    db = next(get_db())
    try:
        if db.query(models.BusinessUnit).count() == 0:
            units = [
                {"name": "หอพัก", "type": models.UnitType.DORMITORY},
                {"name": "อู่ซ่อมรถ", "type": models.UnitType.GARAGE},
                {"name": "บ้านเช่า หลังที่ 1", "type": models.UnitType.HOUSE},
                {"name": "บ้านเช่า หลังที่ 2", "type": models.UnitType.HOUSE},
                {"name": "บ้านเช่า หลังที่ 3", "type": models.UnitType.HOUSE},
            ]
            for unit in units:
                db.add(models.BusinessUnit(**unit))
            db.commit()
            print("Successfully seeded 5 default Business Units.")
    except Exception as e:
        db.rollback()
        print(f"Error seeding business units: {e}")
    finally:
        db.close()
    
    # Seed DormRooms and Houses
    seed_rooms_and_houses()

def seed_rooms_and_houses():
    db = next(get_db())
    try:
        # Seed DormRooms
        if db.query(models.DormRoom).count() == 0:
            rooms = []
            
            # 26_20
            rooms_26_20 = [
                {"number": "401", "rate": 2800, "floor": 4}, 
                {"number": "402", "rate": 2500, "floor": 4},
                {"number": "403", "rate": 2500, "floor": 4}, {"number": "404", "rate": 2200, "floor": 4},
                {"number": "405", "rate": 2500, "floor": 4}, {"number": "406", "rate": 2500, "floor": 4},
                {"number": "301", "rate": 2800, "floor": 3}, 
                {"number": "302", "rate": 2500, "floor": 3},
                {"number": "303", "rate": 2500, "floor": 3}, {"number": "304", "rate": 3000, "floor": 3},
                {"number": "305", "rate": 2800, "floor": 3}, {"number": "306", "rate": 2800, "floor": 3},
                {"number": "307", "rate": 3000, "floor": 3},
                {"number": "201", "rate": 3000, "floor": 2}, {"number": "202", "rate": 2800, "floor": 2},
                {"number": "203", "rate": 2800, "floor": 2}, {"number": "204", "rate": 3000, "floor": 2},
                {"number": "205", "rate": 2800, "floor": 2}, {"number": "206", "rate": 2500, "floor": 2},
                {"number": "207", "rate": 2700, "floor": 2},
                {"number": "101", "rate": 2500, "floor": 1}, {"number": "102", "rate": 2500, "floor": 1},
                {"number": "103", "rate": 2500, "floor": 1}, {"number": "104", "rate": 2500, "floor": 1},
            ]
            for r in rooms_26_20:
                rooms.append(models.DormRoom(
                    dorm_key="26_20",
                    number=r["number"],
                    floor=r["floor"],
                    rate=float(r["rate"]),
                    tenant="",
                    payment_status="vacant",
                    deposit=0.0,
                    lease_status="active"
                ))

            # 26_577
            # Floor 3: 301 to 310
            for i in range(1, 11):
                rooms.append(models.DormRoom(
                    dorm_key="26_577",
                    number=f"3{str(i).zfill(2)}",
                    floor=3,
                    rate=2500.0,
                    tenant="",
                    payment_status="vacant"
                ))
            # Floor 2: 201 to 210
            for i in range(1, 11):
                rooms.append(models.DormRoom(
                    dorm_key="26_577",
                    number=f"2{str(i).zfill(2)}",
                    floor=2,
                    rate=2500.0,
                    tenant="",
                    payment_status="vacant"
                ))
            # Floor 1
            rooms_26_577_f1 = [
                {"number": "101", "rate": 2500, "floor": 1}, 
                {"number": "102", "rate": 2500, "floor": 1},
                {"number": "103", "rate": 2500, "floor": 1}, {"number": "104", "rate": 2000, "floor": 1},
                {"number": "105", "rate": 2500, "floor": 1}, {"number": "106", "rate": 2500, "floor": 1},
                {"number": "107", "rate": 2500, "floor": 1}, {"number": "108", "rate": 2500, "floor": 1},
                {"number": "109", "rate": 2500, "floor": 1}, {"number": "110", "rate": 2500, "floor": 1},
            ]
            for r in rooms_26_577_f1:
                rooms.append(models.DormRoom(
                    dorm_key="26_577",
                    number=r["number"],
                    floor=r["floor"],
                    rate=float(r["rate"]),
                    tenant="",
                    payment_status="vacant",
                    deposit=0.0,
                    lease_status="active"
                ))

            # 73_17
            # Floor 2: B1 to B9
            for i in range(1, 10):
                rooms.append(models.DormRoom(
                    dorm_key="73_17",
                    number=f"B{i}",
                    floor=2,
                    rate=3500.0,
                    tenant="",
                    payment_status="vacant"
                ))
            # Floor 1: A1 to A8
            for i in range(1, 9):
                rooms.append(models.DormRoom(
                    dorm_key="73_17",
                    number=f"A{i}",
                    floor=1,
                    rate=3500.0,
                    tenant="",
                    payment_status="vacant"
                ))

            for r in rooms:
                db.add(r)
            db.commit()
            print(f"Successfully seeded {len(rooms)} Dormitory Rooms.")

        # Seed RentalHouses
        if db.query(models.RentalHouse).count() == 0:
            houses = [
                models.RentalHouse(id="h1", name="บ้านเช่า หลังที่ 1", tenant_name="", monthly_rent=5000.0, payment_status="unpaid", deposit=0.0, lease_status="active"),
                models.RentalHouse(id="h2", name="บ้านเช่า หลังที่ 2", tenant_name="", monthly_rent=4500.0, payment_status="unpaid", deposit=0.0, lease_status="active"),
                models.RentalHouse(id="h3", name="บ้านเช่า หลังที่ 3", tenant_name="", monthly_rent=6000.0, payment_status="unpaid", deposit=0.0, lease_status="active"),
            ]
            for h in houses:
                db.add(h)
            db.commit()
            print("Successfully seeded 3 default Rental Houses.")

    except Exception as e:
        db.rollback()
        print(f"Error seeding rooms and houses: {e}")
    finally:
        db.close()

@app.get("/")
def read_root():
    return {"message": "Welcome to Sovereign Accounting API"}

# Business Units
@app.post("/units/", response_model=schemas.BusinessUnit)
def create_unit(unit: schemas.BusinessUnitCreate, db: Session = Depends(get_db)):
    db_unit = models.BusinessUnit(**unit.dict())
    db.add(db_unit)
    db.commit()
    db.refresh(db_unit)
    return db_unit

@app.get("/units/", response_model=List[schemas.BusinessUnit])
def read_units(db: Session = Depends(get_db)):
    return db.query(models.BusinessUnit).all()

# Customers
@app.post("/customers/", response_model=schemas.Customer)
def create_customer(customer: schemas.CustomerCreate, db: Session = Depends(get_db)):
    db_customer = models.Customer(**customer.dict())
    db.add(db_customer)
    db.commit()
    db.refresh(db_customer)
    return db_customer

@app.get("/customers/", response_model=List[schemas.Customer])
def read_customers(db: Session = Depends(get_db)):
    return db.query(models.Customer).all()

@app.patch("/customers/{customer_id}/", response_model=schemas.Customer)
def update_customer(customer_id: int, customer_update: schemas.CustomerUpdate, db: Session = Depends(get_db)):
    db_customer = db.query(models.Customer).filter(models.Customer.id == customer_id).first()
    if not db_customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    update_data = customer_update.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_customer, key, value)
    
    db.commit()
    db.refresh(db_customer)
    return db_customer

@app.delete("/customers/{customer_id}/")
def delete_customer(customer_id: int, db: Session = Depends(get_db)):
    db_customer = db.query(models.Customer).filter(models.Customer.id == customer_id).first()
    if not db_customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    has_invoices = db.query(models.Invoice).filter(models.Invoice.customer_id == customer_id).first()
    has_tx = db.query(models.Transaction).filter(models.Transaction.customer_id == customer_id).first()
    if has_invoices or has_tx:
        raise HTTPException(status_code=400, detail="ไม่สามารถลบลูกค้าท่านนี้ได้ เนื่องจากมีประวัติธุรกรรมทางการเงินหรือบิลค้างชำระในระบบ (กรุณาใช้การ Soft Delete หรือยกเลิกบิลก่อนเพื่อความถูกต้องทางบัญชี)")
        
    db.delete(db_customer)
    db.commit()
    return {"status": "success", "message": "Customer deleted successfully"}

# Invoices
@app.post("/invoices/", response_model=schemas.Invoice)
def create_invoice(invoice: schemas.InvoiceCreate, db: Session = Depends(get_db)):
    db_invoice = models.Invoice(**invoice.dict())
    db.add(db_invoice)
    db.commit()
    db.refresh(db_invoice)
    return db_invoice

@app.get("/invoices/", response_model=List[schemas.Invoice])
def read_invoices(db: Session = Depends(get_db)):
    return db.query(models.Invoice).all()

@app.patch("/invoices/{invoice_id}/status", response_model=schemas.Invoice)
def update_invoice_status(invoice_id: int, status: str, db: Session = Depends(get_db)):
    db_invoice = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
    if not db_invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    try:
        db_invoice.status = models.InvoiceStatus(status.lower())
        db.commit()
        db.refresh(db_invoice)
        
        ref_id = f"invoice_payment_{db_invoice.id}"
        if status.lower() == "paid":
            existing_tx = db.query(models.Transaction).filter(models.Transaction.reference_id == ref_id).first()
            if not existing_tx:
                db_tx = models.Transaction(
                    type=models.TransactionType.INCOME,
                    amount=db_invoice.amount,
                    description=f"รับชำระบิล: {db_invoice.title} (บิล ID: {db_invoice.id})",
                    reference_id=ref_id,
                    unit_id=db_invoice.unit_id,
                    customer_id=db_invoice.customer_id
                )
                db.add(db_tx)
                db.commit()
        else:
            # Reversal Sync to prevent Ghost Revenue when changing status back from paid
            existing_tx = db.query(models.Transaction).filter(models.Transaction.reference_id == ref_id).first()
            if existing_tx:
                db.delete(existing_tx)
                db.commit()
        
        return db_invoice
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid invoice status")

@app.patch("/invoices/{invoice_id}/", response_model=schemas.Invoice)
def update_invoice(invoice_id: int, invoice_update: schemas.InvoiceUpdate, db: Session = Depends(get_db)):
    db_invoice = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
    if not db_invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Track status and amount changes for Ledger Sync
    old_status = db_invoice.status

    # Update fields
    update_data = invoice_update.dict(exclude_unset=True)
    for key, value in update_data.items():
        if key == "status" and value is not None:
            db_invoice.status = models.InvoiceStatus(value)
        else:
            setattr(db_invoice, key, value)
            
    db.commit()
    db.refresh(db_invoice)
    
    # Ledger Sync Reversal or Update Logic
    ref_id = f"invoice_payment_{db_invoice.id}"
    new_status = db_invoice.status
    
    if new_status == models.InvoiceStatus.PAID:
        existing_tx = db.query(models.Transaction).filter(models.Transaction.reference_id == ref_id).first()
        if not existing_tx:
            # If changed from unpaid/cancelled to paid
            db_tx = models.Transaction(
                type=models.TransactionType.INCOME,
                amount=db_invoice.amount,
                description=f"รับชำระบิล: {db_invoice.title} (บิล ID: {db_invoice.id})",
                reference_id=ref_id,
                unit_id=db_invoice.unit_id,
                customer_id=db_invoice.customer_id
            )
            db.add(db_tx)
            db.commit()
        else:
            # If it was paid and remains paid, but details changed
            existing_tx.amount = db_invoice.amount
            existing_tx.description = f"รับชำระบิล: {db_invoice.title} (บิล ID: {db_invoice.id})"
            existing_tx.unit_id = db_invoice.unit_id
            existing_tx.customer_id = db_invoice.customer_id
            db.commit()
    else:
        # If changed from paid to unpaid/cancelled, delete transaction to prevent Ghost Revenue
        existing_tx = db.query(models.Transaction).filter(models.Transaction.reference_id == ref_id).first()
        if existing_tx:
            db.delete(existing_tx)
            db.commit()

    return db_invoice

@app.delete("/invoices/{invoice_id}/")
def delete_invoice(invoice_id: int, db: Session = Depends(get_db)):
    db_invoice = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
    if not db_invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
        
    # Reversal Sync: delete transaction if it was paid
    ref_id = f"invoice_payment_{invoice_id}"
    existing_tx = db.query(models.Transaction).filter(models.Transaction.reference_id == ref_id).first()
    if existing_tx:
        db.delete(existing_tx)
        
    db.delete(db_invoice)
    db.commit()
    return {"status": "success", "message": "Invoice deleted successfully"}

# Helper to map categories to Thai
EXPENSE_CATEGORY_THAI = {
    models.ExpenseCategory.WATER_BILL: "ค่าน้ำประปาหลวงส่วนกลาง (Water Bill)",
    models.ExpenseCategory.ELECTRIC_BILL: "ค่าไฟฟ้าหลวงส่วนกลาง (Electric Bill)",
    models.ExpenseCategory.SPARE_PARTS: "อะไหล่สำหรับอู่รถ (Spare Parts)",
    models.ExpenseCategory.MAINTENANCE: "ค่าชำระซ่อมบำรุงห้องพัก/บ้านเช่า (Maintenance)",
    models.ExpenseCategory.SALARY: "ค่าแรงช่าง/ค่าจ้างแม่บ้าน (Salary)",
    models.ExpenseCategory.OTHER: "ค่าใช้จ่ายอื่นๆ (Other)"
}

async def send_financial_alert_to_owner(message: str):
    owner_line_user_id = os.getenv("OWNER_LINE_USER_ID")
    if owner_line_user_id:
        try:
            await get_line_bot_api().push_message(PushMessageRequest(
                to=owner_line_user_id,
                messages=[TextMessage(text=message)]
            ))
            return True
        except Exception as e:
            print(f"Error sending owner alert via LINE OA Push: {str(e)}")
            
    token = os.getenv("OWNER_LINE_NOTIFY_TOKEN")
    if not token:
        print("Warning: OWNER_LINE_USER_ID is not configured in .env. (LINE Notify is deprecated as of March 2025; please use OWNER_LINE_USER_ID)")
        return False
        
    url = "https://notify-api.line.me/api/notify"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/x-www-form-urlencoded"
    }
    payload = {"message": message}
    
    import httpx
    try:
        async with httpx.AsyncClient() as client:
            response = await client.post(url, headers=headers, data=payload, timeout=10.0)
        if response.status_code == 200:
            return True
        else:
            print(f"Failed to send LINE Notify: Status {response.status_code}, response: {response.text}")
            return False
    except Exception as e:
        print(f"Error sending LINE Notify: {str(e)}")
        return False

# Transactions
@app.post("/transactions/", response_model=schemas.Transaction)
async def create_transaction(transaction: schemas.TransactionCreate, db: Session = Depends(get_db)):
    db_transaction = models.Transaction(**transaction.dict())
    db.add(db_transaction)
    db.commit()
    db.refresh(db_transaction)
    
    # Trigger LINE Notify for hand-recorded transactions
    time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    amount_formatted = f"{db_transaction.amount:,.2f}"
    
    if db_transaction.type == models.TransactionType.EXPENSE:
        cat_thai = EXPENSE_CATEGORY_THAI.get(db_transaction.expense_category, "ทั่วไป/อื่นๆ (Other)")
        msg = (
            f"\n💸 [เงินออก/บันทึกรายจ่าย] บันทึกรายจ่ายสำเร็จ!\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"🔹 หมวดหมู่: {cat_thai}\n"
            f"🔹 จำนวนเงิน: {amount_formatted} บาท\n"
            f"🔹 คำอธิบาย: {db_transaction.description}\n"
            f"🔹 เวลา: {time_str}"
        )
        await send_financial_alert_to_owner(msg)
    elif db_transaction.type == models.TransactionType.INCOME:
        msg = (
            f"\n💰 [เงินเข้า/บันทึกมือ] บันทึกรายรับด้วยมือสำเร็จ!\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"🔹 จำนวนเงิน: {amount_formatted} บาท\n"
            f"🔹 คำอธิบาย: {db_transaction.description}\n"
            f"🔹 เวลา: {time_str}"
        )
        await send_financial_alert_to_owner(msg)
        
    return db_transaction

@app.get("/transactions/", response_model=List[schemas.Transaction])
def read_transactions(db: Session = Depends(get_db)):
    return db.query(models.Transaction).order_by(models.Transaction.created_at.desc()).all()

@app.patch("/transactions/{transaction_id}/", response_model=schemas.Transaction)
def update_transaction(transaction_id: int, tx_update: schemas.TransactionUpdate, db: Session = Depends(get_db)):
    db_tx = db.query(models.Transaction).filter(models.Transaction.id == transaction_id).first()
    if not db_tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
        
    # Security Rule: Block editing automatic system transactions directly from transactions ledger
    ref = db_tx.reference_id or ""
    if ref.startswith("dorm_payment_") or ref.startswith("house_payment_") or ref.startswith("garage_payment_") or ref.startswith("invoice_payment_"):
        raise HTTPException(status_code=400, detail="ไม่อนุญาตให้แก้ไขธุรกรรมที่เชื่อมต่อระบบอัตโนมัติโดยตรงจากหน้านี้ กรุณาไปแก้ไขหรือเปลี่ยนสถานะที่ Entity ต้นทาง (บิล/อสังหาริมทรัพย์) เพื่อคงความสมบูรณ์และถูกต้องของสมุดบัญชี")
        
    # Update fields
    update_data = tx_update.dict(exclude_unset=True)
    for key, value in update_data.items():
        if key == "type" and value is not None:
            db_tx.type = models.TransactionType(value)
        elif key == "expense_category" and value is not None:
            if value == "":
                db_tx.expense_category = None
            else:
                db_tx.expense_category = models.ExpenseCategory(value)
        else:
            setattr(db_tx, key, value)
            
    db.commit()
    db.refresh(db_tx)
    return db_tx

@app.delete("/transactions/{transaction_id}/")
def delete_transaction(transaction_id: int, db: Session = Depends(get_db)):
    db_tx = db.query(models.Transaction).filter(models.Transaction.id == transaction_id).first()
    if not db_tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
        
    # Security Rule: Block deleting automatic system transactions directly from transactions ledger
    ref = db_tx.reference_id or ""
    if ref.startswith("dorm_payment_") or ref.startswith("house_payment_") or ref.startswith("garage_payment_") or ref.startswith("invoice_payment_"):
        raise HTTPException(status_code=400, detail="ไม่อนุญาตให้ลบหรือโมฆะธุรกรรมที่เชื่อมต่อระบบอัตโนมัติโดยตรงจากหน้านี้ กรุณาไปเปลี่ยนสถานะที่ Entity ต้นทาง (บิล/อสังหาริมทรัพย์) เพื่อคงความถูกต้องของงบแสดงรายรับและยอดสรุปทางการเงิน")
        
    db.delete(db_tx)
    db.commit()
    return {"status": "success", "message": "Transaction deleted successfully"}

from sqlalchemy import func

@app.get("/transactions/summary", response_model=schemas.DashboardSummary)
def get_transaction_summary(db: Session = Depends(get_db)):
    totals = db.query(
        models.Transaction.type,
        func.sum(models.Transaction.amount)
    ).group_by(models.Transaction.type).all()
    
    total_income = next((float(t[1]) for t in totals if t[0] == models.TransactionType.INCOME), 0.0)
    total_expense = next((float(t[1]) for t in totals if t[0] == models.TransactionType.EXPENSE), 0.0)
    balance = total_income - total_expense
    
    unit_totals = db.query(
        models.Transaction.unit_id,
        models.Transaction.type,
        func.sum(models.Transaction.amount)
    ).group_by(models.Transaction.unit_id, models.Transaction.type).all()
    
    units = db.query(models.BusinessUnit).all()
    unit_summaries = []
    
    for u in units:
        u_income = sum(float(t[2]) for t in unit_totals if t[0] == u.id and t[1] == models.TransactionType.INCOME)
        u_expense = sum(float(t[2]) for t in unit_totals if t[0] == u.id and t[1] == models.TransactionType.EXPENSE)
        unit_summaries.append(
            schemas.BusinessUnitSummary(
                id=u.id,
                name=u.name,
                type=u.type,
                total_income=u_income,
                total_expense=u_expense,
                balance=u_income - u_expense
            )
        )
        
    return schemas.DashboardSummary(
        total_income=total_income,
        total_expense=total_expense,
        balance=balance,
        units=unit_summaries
    )

# ==========================================
# Export Excel Endpoint
# ==========================================
@app.get("/transactions/export/excel")
def export_transactions_excel(db: Session = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    transactions = db.query(models.Transaction).order_by(models.Transaction.created_at.desc()).all()
    units = db.query(models.BusinessUnit).all()
    unit_map = {u.id: u.name for u in units}
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "รายการบัญชี"
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ["ลำดับ", "วันที่", "ประเภท", "จำนวนเงิน (บาท)", "รายละเอียด", "ธุรกิจ"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    income_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    expense_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    
    for idx, tx in enumerate(transactions, 1):
        row = idx + 1
        tx_type = "รายรับ" if tx.type == models.TransactionType.INCOME else "รายจ่าย"
        unit_name = unit_map.get(tx.unit_id, "ทั่วไป")
        date_str = tx.created_at.strftime("%d/%m/%Y %H:%M") if tx.created_at else ""
        fill = income_fill if tx.type == models.TransactionType.INCOME else expense_fill
        
        data = [idx, date_str, tx_type, float(tx.amount), tx.description, unit_name]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.fill = fill
            if col == 4:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
    
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)
    
    total_income = sum(float(tx.amount) for tx in transactions if tx.type == models.TransactionType.INCOME)
    total_expense = sum(float(tx.amount) for tx in transactions if tx.type == models.TransactionType.EXPENSE)
    last_row = len(transactions) + 2
    ws.cell(row=last_row, column=3, value="รวมรายรับ:").font = Font(bold=True)
    ws.cell(row=last_row, column=4, value=total_income).font = Font(bold=True, color="228B22")
    ws.cell(row=last_row, column=4).number_format = '#,##0.00'
    ws.cell(row=last_row+1, column=3, value="รวมรายจ่าย:").font = Font(bold=True)
    ws.cell(row=last_row+1, column=4, value=total_expense).font = Font(bold=True, color="FF0000")
    ws.cell(row=last_row+1, column=4).number_format = '#,##0.00'
    ws.cell(row=last_row+2, column=3, value="คงเหลือสุทธิ:").font = Font(bold=True)
    ws.cell(row=last_row+2, column=4, value=total_income - total_expense).font = Font(bold=True, color="1F4E79")
    ws.cell(row=last_row+2, column=4).number_format = '#,##0.00'
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"sovereign_transactions_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/transactions/monthly-summary")
def get_monthly_summary(db: Session = Depends(get_db)):
    six_months_ago = datetime.now() - timedelta(days=180)
    
    transactions = db.query(models.Transaction).filter(
        models.Transaction.created_at >= six_months_ago
    ).all()
    
    monthly_data = {}
    for tx in transactions:
        if tx.created_at:
            key = tx.created_at.strftime("%Y-%m")
            if key not in monthly_data:
                monthly_data[key] = {"month": key, "income": 0.0, "expense": 0.0}
            if tx.type == models.TransactionType.INCOME:
                monthly_data[key]["income"] += float(tx.amount)
            else:
                monthly_data[key]["expense"] += float(tx.amount)
    
    result = sorted(monthly_data.values(), key=lambda x: x["month"])
    return result[-6:] if len(result) > 6 else result

@app.get("/transactions/utility-analytics/", response_model=schemas.UtilityAnalyticsResponse)
def get_utility_margin_analytics(db: Session = Depends(get_db)):
    # Generate last 6 months list
    current_date = datetime.now()
    months_list = []
    for i in range(5, -1, -1):
        m = current_date.month - i
        y = current_date.year
        while m <= 0:
            m += 12
            y -= 1
        months_list.append(f"{y:04d}-{m:02d}")
        
    # Initialize monthly tracking dictionaries
    water_data = {m: {"month": m, "collected": 0.0, "gov_paid": 0.0, "margin": 0.0, "margin_pct": 0.0} for m in months_list}
    elec_data = {m: {"month": m, "collected": 0.0, "gov_paid": 0.0, "margin": 0.0, "margin_pct": 0.0} for m in months_list}
    
    # 1. Fetch and aggregate paid Dorm Payments
    dorm_payments = db.query(models.DormPayment).filter(
        models.DormPayment.payment_status == "paid",
        models.DormPayment.month.in_(months_list)
    ).all()
    for dp in dorm_payments:
        water_data[dp.month]["collected"] += float(dp.water_cost or 0.0)
        elec_data[dp.month]["collected"] += float(dp.electric_cost or 0.0)
        
    # 2. Fetch and aggregate paid House Payments
    house_payments = db.query(models.HousePayment).filter(
        models.HousePayment.payment_status == "paid",
        models.HousePayment.month.in_(months_list)
    ).all()
    for hp in house_payments:
        water_data[hp.month]["collected"] += float(hp.water_bill or 0.0)
        elec_data[hp.month]["collected"] += float(hp.electric_bill or 0.0)
        
    # 3. Fetch and aggregate Transaction expenses for water and electricity bills
    # Note: We filter by Transaction created_at dating within the range of months
    # We pull transactions from the last 200 days to cover our 6 months range
    six_months_ago = current_date - timedelta(days=200)
    transactions = db.query(models.Transaction).filter(
        models.Transaction.type == models.TransactionType.EXPENSE,
        models.Transaction.expense_category.in_([models.ExpenseCategory.WATER_BILL, models.ExpenseCategory.ELECTRIC_BILL]),
        models.Transaction.created_at >= six_months_ago
    ).all()
    
    for tx in transactions:
        if tx.created_at:
            tx_month = tx.created_at.strftime("%Y-%m")
            if tx_month in water_data:
                if tx.expense_category == models.ExpenseCategory.WATER_BILL:
                    water_data[tx_month]["gov_paid"] += float(tx.amount or 0.0)
                elif tx.expense_category == models.ExpenseCategory.ELECTRIC_BILL:
                    elec_data[tx_month]["gov_paid"] += float(tx.amount or 0.0)

    # 4. Calculate Margins and Margin Percentages
    for m in months_list:
        # Water Calculations
        w = water_data[m]
        w["margin"] = w["collected"] - w["gov_paid"]
        if w["gov_paid"] > 0:
            w["margin_pct"] = (w["margin"] / w["gov_paid"]) * 100.0
        else:
            w["margin_pct"] = 100.0 if w["collected"] > 0 else 0.0
            
        # Electricity Calculations
        e = elec_data[m]
        e["margin"] = e["collected"] - e["gov_paid"]
        if e["gov_paid"] > 0:
            e["margin_pct"] = (e["margin"] / e["gov_paid"]) * 100.0
        else:
            e["margin_pct"] = 100.0 if e["collected"] > 0 else 0.0
            
    return {
        "water": [water_data[m] for m in months_list],
        "electricity": [elec_data[m] for m in months_list]
    }

def match_business_unit(description: str, db: Session) -> Optional[models.BusinessUnit]:
    units = db.query(models.BusinessUnit).all()
    if not units:
        return None
    desc_lower = description.lower()
    
    if any(k in desc_lower for k in ["อู่", "ซ่อม", "garage", "รถ", "ตู้"]):
        for u in units:
            if u.type == models.UnitType.GARAGE:
                return u
    if any(k in desc_lower for k in ["บ้าน", "house"]):
        for i in range(1, 4):
            if any(k in desc_lower for k in [str(i), f"หลังที่ {i}", f"หลัง {i}"]):
                for u in units:
                    if f"หลังที่ {i}" in u.name:
                        return u
        for u in units:
            if u.type == models.UnitType.HOUSE:
                return u
    if any(k in desc_lower for k in ["หอ", "dorm", "ห้อง", "b20", "b577", "a20"]):
        for u in units:
            if u.type == models.UnitType.DORMITORY:
                return u
    return None

def find_room_by_line_user_id(line_user_id: str, db: Session) -> Optional[models.DormRoom]:
    # 1. Search in DormRoom.remark
    room = db.query(models.DormRoom).filter(models.DormRoom.remark == line_user_id).first()
    if room:
        return room
    # 2. Search in Customer
    customer = db.query(models.Customer).filter(models.Customer.line_user_id == line_user_id).first()
    if customer:
        # Match Customer.name with DormRoom.tenant
        room = db.query(models.DormRoom).filter(models.DormRoom.tenant == customer.name).first()
        if room:
            return room
    return None

def get_line_user_id_for_room(room: models.DormRoom, db: Session) -> Optional[str]:
    if room.remark and room.remark.startswith("U") and len(room.remark) == 33:
        return room.remark
    if room.tenant:
        customer = db.query(models.Customer).filter(
            models.Customer.name == room.tenant,
            models.Customer.line_user_id.isnot(None),
            models.Customer.line_user_id != ""
        ).first()
        if customer:
            return customer.line_user_id
    return None

@app.post("/webhook")
@app.post("/webhook/")
async def line_webhook(request: Request, db: Session = Depends(get_db)):
    signature = request.headers.get('X-Line-Signature')
    body = await request.body()
    body_str = body.decode('utf-8')

    try:
        events = parser.parse(body_str, signature)
    except InvalidSignatureError:
        raise HTTPException(status_code=400, detail="Invalid signature")

    for event in events:
        if not isinstance(event, MessageEvent):
            continue
        
        user_id = event.source.user_id

        # 📸 Handle Image/Slip uploads sent directly to chat
        if event.message.type == "image":
            # 1. Try to find the associated room for the user
            room = find_room_by_line_user_id(user_id, db)
            
            if not room:
                reply_text = (
                    "❌ ขออภัยค่ะ ระบบยังไม่สามารถตรวจสอบสลิปนี้ให้โดยอัตโนมัติได้ เนื่องจากไลน์นี้ยังไม่ได้ผูกบัญชีห้องพักในระบบค่ะ\n\n"
                    "กรุณาผูกบัญชีโดยส่งข้อความในรูปแบบนี้ก่อนนะคะ:\n"
                    "👉 [ชื่อเล่น] หอ [เลขหอ] ห้อง [เลขห้อง]\n"
                    "ตัวอย่าง: แก้ว หอ 26/20 ห้อง 302\n\n"
                    "หรือติดต่อเจ้าหน้าที่ผู้ดูแลระบบเพื่อทำการผูกบัญชีให้ได้เลยค่ะ 🙏😊"
                )
                await get_line_bot_api().reply_message(
                    ReplyMessageRequest(
                        reply_token=event.reply_token,
                        messages=[TextMessage(text=reply_text)]
                    )
                )
                continue
                
            # Find customer details for this user
            customer = db.query(models.Customer).filter(models.Customer.line_user_id == user_id).first()
            if not customer and room.tenant:
                customer = db.query(models.Customer).filter(models.Customer.name == room.tenant).first()
                
            # Gather all unpaid bills/invoices to match dynamically
            unpaid_bills = []
            
            if room.payment_status != "paid":
                room_total = room.rate + room.water_cost + room.electric_cost + room.cleaning_fee + room.other_fee + room.fine_cost
                unpaid_bills.append({
                    "type": "dorm",
                    "amount": float(room_total),
                    "title": f"ค่าห้องพักประจำเดือน (ห้อง {room.number})",
                    "obj": room
                })
                
            if customer:
                invoices = db.query(models.Invoice).filter(
                    models.Invoice.customer_id == customer.id,
                    models.Invoice.status == models.InvoiceStatus.UNPAID
                ).all()
                for inv in invoices:
                    unpaid_bills.append({
                        "type": "invoice",
                        "amount": float(inv.amount),
                        "title": f"บิลพิเศษ: {inv.title}",
                        "obj": inv
                    })
                    
            if not unpaid_bills:
                reply_text = (
                    f"💡 ระบบพบว่าคุณไม่มียอดค้างชำระในขณะนี้ค่ะ!\n"
                    f"ห้องพัก หอ {room.dorm_key.replace('_', '/')} ห้อง {room.number} ของคุณ ได้รับการชำระเรียบร้อยแล้วค่ะ\n\n"
                    f"หากรูปภาพนี้เป็นหลักฐานสลิปสำหรับค่าใช้จ่ายอื่น หรือยอดเพิ่มเติมนอกระบบ เจ้าหน้าที่แอดมินจะตรวจสอบสลิปและดำเนินการบันทึกเข้าระบบบัญชีให้ภายหลังนะคะ ขอบพระคุณค่ะ 🙏😊"
                )
                await get_line_bot_api().reply_message(
                    ReplyMessageRequest(
                        reply_token=event.reply_token,
                        messages=[TextMessage(text=reply_text)]
                    )
                )
                continue

            # 2. Proceed with download and SlipOK API validation
            try:
                # Retrieve the image binary content from LINE
                blob_api = get_line_blob_api()
                message_content = await blob_api.get_message_content(message_id=event.message.id)
                img_bytes = bytes(message_content)
                
                slipok_api_key = os.getenv("SLIPOK_API_KEY")
                slipok_branch_id = os.getenv("SLIPOK_BRANCH_ID", "0")
                
                import httpx
                import random
                
                matched_bill = None
                amount_paid = 0.0
                ref_no = ""
                sender_name = room.tenant or "ผู้เช่า"
                receiver_name = "บริษัท Sovereign Volt จำกัด"
                trans_date_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                is_sandbox = True
                
                if slipok_api_key and slipok_api_key != "your_slipok_api_key_here":
                    # Production Mode: Call SlipOK API using files upload (no amount parameter in payload for dynamic matching)
                    is_sandbox = False
                    url = f"https://api.slipok.com/api/line/apikey/{slipok_branch_id}"
                    headers = {
                        "x-authorization": slipok_api_key
                    }
                    files = {
                        "files": ("slip.jpg", img_bytes, "image/jpeg")
                    }
                    payload = {
                        "log": "true"
                    }
                    
                    async with httpx.AsyncClient() as client:
                        response = await client.post(url, headers=headers, files=files, data=payload, timeout=20.0)
                        
                    if response.status_code != 200:
                        raise Exception(f"SlipOK API Error (HTTP {response.status_code}): {response.text}")
                        
                    res_json = response.json()
                    success = res_json.get("success", False)
                    data = res_json.get("data", {})
                    is_valid = success and (data.get("success", True) if isinstance(data, dict) else True)
                    
                    if not is_valid:
                        error_msg = res_json.get("message") or (data.get("message") if isinstance(data, dict) else None) or "สลิปไม่ผ่านการตรวจสอบความถูกต้อง"
                        raise Exception(error_msg)
                        
                    ref_no = data.get("transRef")
                    if not ref_no:
                        raise Exception("ไม่พบรหัสอ้างอิงธุรกรรม (Transaction Reference) ในสลิป")
                        
                    # Check for duplicates
                    existing_tx = db.query(models.Transaction).filter(models.Transaction.reference_id == ref_no).first()
                    if existing_tx:
                        reply_text = (
                            "❌ ระบบไม่สามารถยืนยันสลิปนี้ได้ เนื่องจากสลิปโอนเงินนี้เคยถูกส่งเพื่อยืนยันยอดเงินไปแล้วในระบบ เพื่อความปลอดภัยป้องกันการตรวจจับซ้ำ กรุณาตรวจสอบหรือติดต่อแอดมินโดยตรงค่ะ"
                        )
                        await get_line_bot_api().reply_message(
                            ReplyMessageRequest(reply_token=event.reply_token, messages=[TextMessage(text=reply_text)])
                        )
                        continue
                        
                    sender_name = data.get("sender", {}).get("displayName") or data.get("sender", {}).get("name") or room.tenant or "ผู้เช่าในระบบ"
                    receiver_name = data.get("receiver", {}).get("displayName") or data.get("receiver", {}).get("name") or "หอพัก/บริษัท"
                    amount_paid = float(data.get("amount", 0.0))
                    trans_date_str = data.get("transDate") or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    
                    # Search and match the amount from slip against outstanding bills
                    for bill in unpaid_bills:
                        if abs(amount_paid - bill["amount"]) <= 0.01:
                            matched_bill = bill
                            break
                            
                    if not matched_bill:
                        # Incorrect amount or doesn't match any bill
                        unpaid_details = "\n".join([f"- {b['title']}: {b['amount']:,.2f} บาท" for b in unpaid_bills])
                        reply_text = (
                            f"❌ ระบบสแกนพบลดสลิปโอนเงินจริงที่ยอด {amount_paid:,.2f} บาท\n"
                            f"แต่ยอดเงินดังกล่าวไม่สอดคล้องกับรายการค้างชำระใบใดเลยของคุณในระบบค่ะ\n\n"
                            f"📋 รายการที่คุณค้างชำระในขณะนี้:\n{unpaid_details}\n\n"
                            f"กรุณาโอนเงินให้ตรงตามยอดบิล หรือติดต่อแอดมินโดยตรงเพื่อปรับปรุงยอดค่ะ 🙏😊"
                        )
                        await get_line_bot_api().reply_message(
                            ReplyMessageRequest(reply_token=event.reply_token, messages=[TextMessage(text=reply_text)])
                        )
                        continue
                else:
                    # Sandbox Fallback mode (simulated success)
                    # For sandbox, if there's only 1 unpaid bill, match it. 
                    # If there are multiple, match the first one (room bill preferred) and display sandbox warning.
                    matched_bill = unpaid_bills[0]
                    amount_paid = matched_bill["amount"]
                    
                    ref_no = f"SIM-SLIP-{datetime.now().strftime('%Y%m%d')}-" + "".join([str(random.randint(0, 9)) for _ in range(6)])
                    while db.query(models.Transaction).filter(models.Transaction.reference_id == ref_no).first():
                        ref_no = f"SIM-SLIP-{datetime.now().strftime('%Y%m%d')}-" + "".join([str(random.randint(0, 9)) for _ in range(6)])
                        
                    sender_name = room.tenant or "ผู้เช่าจำลอง (โหมดทดสอบ)"
                    receiver_name = "บริษัท หอพัก Sovereign (โหมดทดสอบ)"
                    trans_date_str = datetime.now().strftime("%d/%m/%Y %H:%M")
                    
                # Bookkeeping & Saving payment status
                current_month_key = get_current_billing_month(trans_date_str)
                payment_type = matched_bill["type"]
                expected_amount = matched_bill["amount"]
                target_obj = matched_bill["obj"]
                
                if payment_type == "dorm":
                    target_obj.payment_status = "paid"
                    target_obj.payment_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    
                    db.add(models.Transaction(
                        type=models.TransactionType.INCOME,
                        amount=expected_amount,
                        description=f"ค่าเช่าห้อง {target_obj.number} รอบ {current_month_key} (ผู้โอน: {sender_name}, อ้างอิงสลิป: {ref_no})",
                        reference_id=ref_no,
                        unit_id=(lambda u: u.id if u else None)(db.query(models.BusinessUnit).filter(models.BusinessUnit.type == models.UnitType.DORMITORY).first())
                    ))
                    
                    # Add DormPayment if not already exists for the month
                    if not db.query(models.DormPayment).filter(models.DormPayment.room_id == target_obj.id, models.DormPayment.month == current_month_key).first():
                        db.add(models.DormPayment(
                            room_id=target_obj.id,
                            month=current_month_key,
                            amount=target_obj.rate,
                            water_cost=target_obj.water_cost,
                            electric_cost=target_obj.electric_cost,
                            cleaning_fee=target_obj.cleaning_fee,
                            other_fee=target_obj.other_fee,
                            fine_cost=target_obj.fine_cost,
                            payment_status="paid",
                            paid_at=datetime.utcnow()
                        ))
                    
                else:
                    # Invoice type
                    target_obj.status = models.InvoiceStatus.PAID
                    db.add(models.Transaction(
                        type=models.TransactionType.INCOME,
                        amount=expected_amount,
                        description=f"ชำระบิลเรียกเก็บเงิน: {target_obj.title} #{target_obj.id} (ผู้โอน: {sender_name}, อ้างอิงสลิป: {ref_no})",
                        reference_id=ref_no,
                        unit_id=target_obj.unit_id,
                        customer_id=target_obj.customer_id
                    ))
                    
                db.commit()
                
                # Notify Owner about the successful payment
                notify_msg = (
                    f"🔔 ยืนยันยอดชำระเงินอัตโนมัติ (Dynamic Match)! 🔔\n"
                    f"━━━━━━━━━━━━━━━━━━\n"
                    f"🚪 หอพัก/ห้อง: ห้อง {room.number}\n"
                    f"👤 ผู้เช่า: {room.tenant}\n"
                    f"🧾 บิลที่ชำระ: {matched_bill['title']}\n"
                    f"💰 ยอดเงินชำระ: {expected_amount:,.2f} บาท\n"
                    f"👤 ผู้โอน: {sender_name}\n"
                    f"🔢 รหัสอ้างอิง: {ref_no}\n"
                    f"━━━━━━━━━━━━━━━━━━\n"
                    f"ระบบได้จับคู่ยอดอัตโนมัติ บันทึกบัญชี และออกใบเสร็จเรียบร้อยแล้วค่ะ ✨"
                )
                await send_financial_alert_to_owner(notify_msg)
                
                # Reply to Tenant with template
                sandbox_note = "\n⚠️ (โหมดทดสอบ Sandbox) ระบบทำการจับคู่จำลองบิลค้างชำระให้อัตโนมัติ" if is_sandbox and len(unpaid_bills) > 1 else ""
                
                reply_text = (
                    f"🎉 ระบบจับคู่และชำระบิลอัตโนมัติสำเร็จ!\n"
                    f"━━━━━━━━━━━━━━━━━━\n"
                    f"🧾 รายการ: {matched_bill['title']}\n"
                    f"💰 ยอดเงิน: {amount_paid:,.2f} บาท\n"
                    f"📅 วันเวลาโอน: {trans_date_str}\n"
                    f"🔢 รหัสสลิป: {ref_no}\n"
                    f"━━━━━━━━━━━━━━━━━━\n"
                    f"ยอดเงินของคุณได้รับการชำระและตัดยอดออกจากบิลเรียบร้อยแล้วค่ะ ขอบพระคุณค่ะ 🙏😊{sandbox_note}"
                )
                await get_line_bot_api().reply_message(
                    ReplyMessageRequest(
                        reply_token=event.reply_token,
                        messages=[TextMessage(text=reply_text)]
                    )
                )
                
            except Exception as e:
                db.rollback()
                reply_text = (
                    f"❌ ระบบไม่สามารถยืนยันสลิปนี้ได้ เนื่องจากเกิดข้อผิดพลาดในการตรวจสอบหรือบันทึกบัญชีค่ะ\n\n"
                    f"รายละเอียดข้อผิดพลาด: {str(e)}\n\n"
                    f"กรุณาตรวจสอบความถูกต้องของสลิป หรือติดต่อส่งข้อมูลให้แอดมินโดยตรงเพื่อช่วยเหลือค่ะ 🙏😊"
                )
                await get_line_bot_api().reply_message(
                    ReplyMessageRequest(
                        reply_token=event.reply_token,
                        messages=[TextMessage(text=reply_text)]
                    )
                )
            
            continue

        if not isinstance(event.message, TextMessageContent):
            continue
        
        text = event.message.text.strip()

        if text == "แจ้งซ่อม":
            reply_text = (
                "🔧 ระบบรับแจ้งเตือนปัญหาซ่อมแซมห้องพักอัตโนมัติ:\n\n"
                "กรุณาพิมพ์รายละเอียดปัญหาโดยระบุเลขห้องของท่านตามฟอร์แมตด้านล่างนี้ได้เลยค่ะ:\n"
                "👉 \"แจ้งซ่อม หอ[เลขหอ] ห้อง [เลขห้อง] [ระบุรายละเอียดของปัญหาชำรุด]\"\n"
                "👉ขอความร่วมมือส่งรูป เเละวิดีโอมาด้วยค่ะ\n\n"
                "ตัวอย่างเช่น:\n"
                "💬 แจ้งซ่อม หอ 26/577 ห้อง 302 ลูกบิดประตูด้านในฝืดและปลดล็อกไม่ได้\n"
                "💬 แจ้งซ่อม หอ 26/20 ห้อง B5 หลอดไฟในห้องน้ำดับและต้องการให้เปลี่ยนหลอดไฟใหม่\n\n"
                "เมื่อส่งข้อมูลแล้ว ระบบจะยิงแจ้งเตือนหาเจ้าของหอพักทันที ✨\n\n"
                "!!!หากมีเหตุเร่งด่วน!!!\n"
                "โทร: 081-933-0490"
            )
            await get_line_bot_api().reply_message(ReplyMessageRequest(reply_token=event.reply_token, messages=[TextMessage(text=reply_text)]))
            continue

        if text == "แจ้งโอนเงิน":
            reply_text = (
                "💵 ช่องทางการชำระเงินของหอพัก:\n"
                "ท่านสามารถชำระเงินผ่านบัญชี PromptPay หรือโอนเข้าบัญชีธนาคารของหอพัก (ข้อมูลแสดงอยู่ในใบแจ้งหนี้)\n\n"
                "เมื่อโอนเงินเสร็จเรียบร้อยแล้ว:\n"
                "👉 กรุณาส่งภาพสลิปโอนเงินเข้ามาในห้องแชทนี้ได้ทันทีค่ะ/ครับ\n\n"
                "⚠️ ข้อแนะนำ: ระบบหลังบ้านของเราติดตั้งระบบ SlipOK AI ในการตรวจสลิปปลอมและตรวจหาการเคลมยอดเงินซ้ำ หากส่งสลิปแล้วกรุณารอระบบประมวลผลสักครู่ค่ะ\n\n"
                "ได้รับข้อความเเล้ว(LINE messenger) รอสักครู่น้าา(shiny)\n\n"
                "(warning)หากอุปกรณ์ภายในห้องชำรุด(warning)\n"
                "ขอความร่วมมือส่งรูป เเละวิดีโอมาด้วยค่ะ(incoming call)\n\n"
                "หากมีเหตุเร่งด่วน(oh no!)(!)\n"
                "โทร: 081-933-0490 หรือ 093-130-5336"
            )
            await get_line_bot_api().reply_message(ReplyMessageRequest(reply_token=event.reply_token, messages=[TextMessage(text=reply_text)]))
            continue

        if text.startswith("แจ้งซ่อม"):
            content = text[len("แจ้งซ่อม"):].strip()
            
            # Smart Regex: extract optional dorm and room number
            dorm_match = re.search(r'หอ\s*(26/20|26/577|73/17|26_20|26_577|73_17|26-20|26-577|73-17)', content, re.IGNORECASE)
            room_match = re.search(r'ห้อง\s*([a-zA-Z0-9_\-/]+)', content, re.IGNORECASE)
            
            room = None
            description = content
            
            if room_match:
                room_candidate = room_match.group(1).strip()
                dorm_candidate = None
                if dorm_match:
                    dorm_input = dorm_match.group(1).strip()
                    dorm_candidate = dorm_input.replace("/", "_").replace("-", "_")
                
                if dorm_candidate:
                    # Query matching BOTH dorm key and room number
                    room = db.query(models.DormRoom).filter(
                        models.DormRoom.dorm_key == dorm_candidate,
                        models.DormRoom.number == room_candidate
                    ).first()
                else:
                    # Query matching room number only
                    room = db.query(models.DormRoom).filter(models.DormRoom.number == room_candidate).first()
                
                # Clean up description text (strip out matched หอ and ห้อง patterns)
                desc_clean = content
                if dorm_match:
                    desc_clean = desc_clean.replace(dorm_match.group(0), "", 1)
                if room_match:
                    desc_clean = desc_clean.replace(room_match.group(0), "", 1)
                
                # Strip out residual "หอ" or "ห้อง" text if any
                desc_clean = re.sub(r'^\s*หอ\s*ห้อง\s*', '', desc_clean)
                desc_clean = re.sub(r'^[\s\-:]+', '', desc_clean).strip()
                description = desc_clean
            
            # If no room matched from text, try auto-bound fallback using LINE User ID
            if not room:
                room = find_room_by_line_user_id(user_id, db)
                
            if not room:
                reply_text = (
                    "❌ ขออภัยค่ะ ระบบไม่พบข้อมูลเลขห้องของท่าน\n\n"
                    "กรุณาแจ้งซ่อมตามรูปแบบนี้นะคะ:\n"
                    "👉 แจ้งซ่อม หอ [เลขหอ] ห้อง [เลขห้อง] [ปัญหาที่พบ]\n"
                    "ตัวอย่าง: แจ้งซ่อม หอ 26/20 ห้อง 302 ท่อน้ำทิ้งอุดตัน\n\n"
                    "หรือติดต่อเจ้าหน้าที่เพื่อผูกไอดี LINE กับห้องพักของท่านค่ะ"
                )
            else:
                ticket_desc = description or "ไม่ระบุรายละเอียดปัญหา"
                ticket = models.MaintenanceTicket(
                    room_number=room.number,
                    description=ticket_desc,
                    status="pending",
                    line_user_id=user_id,
                    created_at=datetime.utcnow()
                )
                db.add(ticket)
                db.commit()
                db.refresh(ticket)
                
                alert_message = (
                    f"\n🚨 แจ้งซ่อมปัญหาหอพัก! 🚨\n"
                    f"━━━━━━━━━━━━━━━━━━\n"
                    f"🏢 ห้องพัก: ห้อง {room.number}\n"
                    f"👤 ผู้แจ้ง: {room.tenant or 'ไม่ระบุชื่อ'}\n"
                    f"🔧 รายละเอียด: {ticket_desc}\n"
                    f"📅 วันเวลา: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
                    f"━━━━━━━━━━━━━━━━━━\n"
                    f"🔗 จัดการใบสั่งงานผ่านหน้าเว็บ ERP"
                )
                await send_financial_alert_to_owner(alert_message)
                
                reply_text = (
                    f"📝 บันทึกใบงานแจ้งซ่อมเรียบร้อยแล้วครับ!\n"
                    f"━━━━━━━━━━━━━━━━━━\n"
                    f"🎫 ใบรับงานเลขที่: #{ticket.id}\n"
                    f"🏢 ห้องพัก: ห้อง {ticket.room_number}\n"
                    f"🔧 อาการชำรุด: {ticket.description}\n"
                    f"⚙️ สถานะ: รอดำเนินการ (Pending)\n"
                    f"━━━━━━━━━━━━━━━━━━\n"
                    f"ระบบได้ส่งการแจ้งเตือนไปยังผู้ดูแลระบบและช่างเรียบร้อยแล้ว เราจะเร่งดำเนินการตรวจสอบให้โดยเร็วที่สุดครับ ขอบคุณครับ 🙏"
                )
                
            await get_line_bot_api().reply_message(ReplyMessageRequest(reply_token=event.reply_token, messages=[TextMessage(text=reply_text)]))
            continue

        # 👤 Auto-binding pattern: [Nickname] หอ [DormKey] ห้อง [RoomNumber]
        # e.g., "แก้ว หอ 26/20 ห้อง 302"
        binding_match = re.search(
            r'^([^\n]+?)\s*หอ\s*(26/20|26/577|73/17|26_20|26_577|73_17|26-20|26-577|73-17)\s*ห้อง\s*([a-zA-Z0-9_\-/]+)',
            text,
            re.IGNORECASE
        )
        if binding_match:
            nickname = binding_match.group(1).strip()
            dorm_input = binding_match.group(2).strip()
            room_number = binding_match.group(3).strip()

            # Normalize dorm_input (e.g. 26/20 -> 26_20)
            dorm_key = dorm_input.replace("/", "_").replace("-", "_")

            # Find the Room
            room = db.query(models.DormRoom).filter(
                models.DormRoom.dorm_key == dorm_key,
                models.DormRoom.number == room_number
            ).first()

            if not room:
                reply_text = (
                    f"❌ ขออภัยค่ะ ไม่พบข้อมูลห้องพักหมายเลข {room_number} ในหอพัก {dorm_input} ในระบบค่ะ\n\n"
                    f"กรุณาตรวจสอบเลขห้องและชื่อหอพักอีกครั้ง หรือพิมพ์แจ้งผู้ดูแลระบบเพื่อทำการแก้ไขข้อมูลนะคะ 🙏"
                )
            else:
                # 1. Clear this line_user_id from other rooms (avoid duplicates/stale links)
                db.query(models.DormRoom).filter(
                    models.DormRoom.remark == user_id,
                    models.DormRoom.id != room.id
                ).update({models.DormRoom.remark: None}, synchronize_session=False)

                # 2. Clear this line_user_id from other customers to maintain unique constraint
                db.query(models.Customer).filter(
                    models.Customer.line_user_id == user_id
                ).update({models.Customer.line_user_id: None}, synchronize_session=False)

                # 3. Bind the Line User ID to the Room
                room.remark = user_id

                # 4. Bind or Create Customer under Dormitory unit
                dorm_unit = db.query(models.BusinessUnit).filter(models.BusinessUnit.type == models.UnitType.DORMITORY).first()
                
                # If room tenant is empty, populate it with the nickname
                tenant_name = room.tenant or nickname
                if not room.tenant:
                    room.tenant = tenant_name

                # Look up customer
                customer = None
                if dorm_unit:
                    customer = db.query(models.Customer).filter(
                        models.Customer.name == tenant_name,
                        models.Customer.unit_id == dorm_unit.id
                    ).first()

                if customer:
                    customer.line_user_id = user_id
                else:
                    customer = models.Customer(
                        name=tenant_name,
                        line_user_id=user_id,
                        unit_id=dorm_unit.id if dorm_unit else None
                    )
                    db.add(customer)

                db.commit()

                reply_text = (
                    f"🏠✨ ผูกบัญชีห้องพักสำเร็จเรียบร้อยแล้วค่ะ!\n"
                    f"━━━━━━━━━━━━━━━━━━\n"
                    f"👤 ผู้เช่า: คุณ {tenant_name}\n"
                    f"🏢 หอพัก: หอ {dorm_input}\n"
                    f"🚪 ห้องพัก: ห้อง {room_number}\n"
                    f"━━━━━━━━━━━━━━━━━━\n"
                    f"ยินดีต้อนรับเข้าสู่ระบบ LINE OA สำหรับลูกหอพักค่ะ 🏠✨\n\n"
                    f"ท่านสามารถใช้แผง Rich Menu ด้านล่าง เพื่อความสะดวกในการใช้งานดังนี้:\n"
                    f"1. 🧾 [เช็คยอด] เพื่อตรวจสอบบิลค่าเช่า/ค่าน้ำ/ค่าไฟ รายเดือน\n"
                    f"2. 💵 [แจ้งโอนเงิน] ส่งภาพสลิปโอนเงิน PromptPay เพื่อให้ระบบ AI ตรวจสอบยอดอัตโนมัติ\n"
                    f"3. 🔧 [แจ้งซ่อม] พิมพ์แจ้งปัญหาความชำรุดในห้องพักเพื่อให้ช่างเข้าไปดำเนินการแก้ไข\n\n"
                    f"ขอบคุณที่ใช้บริการค่ะ 🙏😊"
                )

            await get_line_bot_api().reply_message(
                ReplyMessageRequest(
                    reply_token=event.reply_token,
                    messages=[TextMessage(text=reply_text)]
                )
            )
            continue



        if text == "เช็คยอด":
            # 1. Search for customer by LINE ID
            customer = db.query(models.Customer).filter(models.Customer.line_user_id == user_id).first()
            
            # 2. Search for room by LINE ID or customer name match
            room = find_room_by_line_user_id(user_id, db)
            if not customer and room and room.tenant:
                customer = db.query(models.Customer).filter(models.Customer.name == room.tenant).first()
                
            if not customer and not room:
                reply_message = TextMessage(text="❌ ขออภัยค่ะ ไม่พบข้อมูลห้องพักหรือข้อมูลของคุณในระบบ กรุณาติดต่อแอดมินเพื่อทำการผูกบัญชี LINE ค่ะ 🙏😊")
            else:
                tenant_name = customer.name if customer else (room.tenant if room else "ผู้เช่า")
                unpaid_items = []
                
                # Check for unpaid Room Bill
                if room and room.payment_status != "paid":
                    rate_val = room.rate or 0.0
                    water_val = room.water_cost or 0.0
                    electric_val = room.electric_cost or 0.0
                    cleaning_val = room.cleaning_fee or 0.0
                    other_val = room.other_fee or 0.0
                    fine_val = room.fine_cost or 0.0
                    
                    room_total = rate_val + water_val + electric_val + cleaning_val + other_val + fine_val
                    unpaid_items.append({
                        "type": "room",
                        "title": f"🚪 ค่าห้องพักประจำเดือน (ห้อง {room.number})",
                        "amount": room_total,
                        "details": [
                            f"• ค่าเช่าห้อง: {rate_val:,.2f} บาท",
                            f"• ค่าน้ำประปา: {water_val:,.2f} บาท",
                            f"• ค่าไฟฟ้า: {electric_val:,.2f} บาท",
                            f"• ค่าบริการอื่นๆ/ค่าปรับ: {(cleaning_val + other_val + fine_val):,.2f} บาท"
                        ]
                    })
                
                # Check for other unpaid Custom Invoices
                if customer:
                    unpaid_invoices = db.query(models.Invoice).filter(
                        models.Invoice.customer_id == customer.id,
                        models.Invoice.status == models.InvoiceStatus.UNPAID
                    ).all()
                    for inv in unpaid_invoices:
                        unpaid_items.append({
                            "type": "invoice",
                            "title": f"🧾 บิลพิเศษ: {inv.title}",
                            "amount": inv.amount,
                            "details": [
                                f"• เลขที่บิล: #{inv.id}",
                                f"• รายละเอียด: {inv.title}"
                            ]
                        })
                
                if not unpaid_items:
                    reply_text = (
                        f"สวัสดีค่ะ คุณ {tenant_name} 🏠✨\n\n"
                        f"🎉 ยินดีด้วยค่ะ! คุณไม่มียอดค้างชำระในระบบขณะนี้ค่ะ\n"
                        f"ขอบพระคุณที่เป็นลูกบ้านที่ดีเสมอมานะคะ รักษาสุขภาพด้วยค่ะ 🙏😊"
                    )
                    reply_message = TextMessage(text=reply_text)
                else:
                    import json
                    if len(unpaid_items) == 1 and unpaid_items[0]["type"] == "room" and room:
                        current_month_key = get_current_billing_month()
                        flex_payload = create_dorm_bill_flex(room, current_month_key)
                        reply_message = FlexMessage(
                            alt_text=f"ใบแจ้งยอดค่าเช่าห้อง {room.number} ประจำรอบบิล {current_month_key}",
                            contents=FlexContainer.from_json(json.dumps(flex_payload))
                        )
                    else:
                        grand_total = sum(item["amount"] for item in unpaid_items)
                        room_number = room.number if room else "-"
                        flex_payload = create_consolidated_bill_flex(tenant_name, room_number, unpaid_items, grand_total)
                        reply_message = FlexMessage(
                            alt_text=f"สรุปยอดค้างชำระทั้งหมดของคุณ {tenant_name}",
                            contents=FlexContainer.from_json(json.dumps(flex_payload))
                        )
                    
            await get_line_bot_api().reply_message(
                ReplyMessageRequest(
                    reply_token=event.reply_token,
                    messages=[reply_message]
                )
            )
            continue

        # 🛡️ Admin Security Check
        admin_commands = ["สรุปวันนี้", "ยอดเงิน", "ทวงบิล"]
        is_admin_cmd = text in admin_commands or bool(re.match(r'^(รับ|จ่าย)\s+(\d+(?:\.\d+)?)\s+(.*)$', text, re.IGNORECASE))
        
        if is_admin_cmd:
            owner_line_id = os.getenv("OWNER_LINE_USER_ID")
            if not owner_line_id or user_id != owner_line_id:
                # Silently ignore to prevent normal tenants from discovering admin commands
                continue

        if text == "สรุปวันนี้":
            today_start = datetime.combine(date.today(), time.min)
            today_end = datetime.combine(date.today(), time.max)
            today_tx = db.query(models.Transaction).filter(models.Transaction.created_at >= today_start, models.Transaction.created_at <= today_end).all()
            total_inc = sum(t.amount for t in today_tx if t.type == models.TransactionType.INCOME)
            total_exp = sum(t.amount for t in today_tx if t.type == models.TransactionType.EXPENSE)
            reply_text = f"📅 สรุปยอดบัญชีวันนี้\n📈 รายรับรวม: +{total_inc:,.2f} บาท\n📉 รายจ่ายรวม: -{total_exp:,.2f} บาท\n💰 คงเหลือสุทธิ: {total_inc - total_exp:,.2f} บาท"
            await get_line_bot_api().reply_message(ReplyMessageRequest(reply_token=event.reply_token, messages=[TextMessage(text=reply_text)]))
            continue

        if text == "ยอดเงิน":
            units = db.query(models.BusinessUnit).all()
            transactions = db.query(models.Transaction).all()
            total_inc_all = sum(t.amount for t in transactions if t.type == models.TransactionType.INCOME)
            total_exp_all = sum(t.amount for t in transactions if t.type == models.TransactionType.EXPENSE)
            lines = [f"🏢 {u.name}: {(sum(t.amount for t in transactions if t.unit_id == u.id and t.type == models.TransactionType.INCOME) - sum(t.amount for t in transactions if t.unit_id == u.id and t.type == models.TransactionType.EXPENSE)):,.2f} บาท" for u in units]
            reply_text = f"💰 ยอดเงินคงเหลือแต่ละธุรกิจ\n{chr(10).join(lines)}\n💵 ยอดรวมทั้งหมด: {total_inc_all - total_exp_all:,.2f} บาท"
            await get_line_bot_api().reply_message(ReplyMessageRequest(reply_token=event.reply_token, messages=[TextMessage(text=reply_text)]))
            continue

        if text == "ทวงบิล":
            unpaid_rooms = db.query(models.DormRoom).filter(models.DormRoom.payment_status != "paid", models.DormRoom.tenant != "", models.DormRoom.tenant != None).all()
            unpaid_houses = db.query(models.RentalHouse).filter(models.RentalHouse.payment_status != "paid", models.RentalHouse.tenant_name != "", models.RentalHouse.tenant_name != None).all()
            unpaid_jobs = db.query(models.GarageJob).filter(models.GarageJob.payment_status != "paid").all()
            lines = ["🔔 สรุปบิลค้างชำระ"]
            if unpaid_rooms:
                lines.append(f"\n🏢 หอพัก ({len(unpaid_rooms)} ห้อง):")
                for r in unpaid_rooms[:15]: lines.append(f"  ห้อง {r.number} - {r.tenant}: {(r.rate + r.water_cost + r.electric_cost + r.cleaning_fee + r.other_fee + r.fine_cost):,.0f}฿")
            if unpaid_houses:
                lines.append(f"\n🏠 บ้านเช่า ({len(unpaid_houses)} หลัง):")
                for h in unpaid_houses: lines.append(f"  {h.name} - {h.tenant_name}: {(h.monthly_rent + h.water_bill + h.electric_bill):,.0f}฿")
            if unpaid_jobs:
                lines.append(f"\n🔧 อู่ซ่อมรถ ({len(unpaid_jobs)} งาน):")
                for j in unpaid_jobs[:10]: lines.append(f"  {j.license_plate} ({j.customer_name}): {j.total_cost:,.0f}฿")
            if not (unpaid_rooms or unpaid_houses or unpaid_jobs): lines.append("\n✅ ไม่มีบิลค้างชำระครับ!")
            await get_line_bot_api().reply_message(ReplyMessageRequest(reply_token=event.reply_token, messages=[TextMessage(text=chr(10).join(lines))]))
            continue

        match = re.match(r'^(รับ|จ่าย)\s+(\d+(?:\.\d+)?)\s+(.*)$', text, re.IGNORECASE)
        if match:
            action, amount, description = match.group(1).lower(), float(match.group(2)), match.group(3).strip()
            tx_type = models.TransactionType.INCOME if action == "รับ" else models.TransactionType.EXPENSE
            unit = match_business_unit(description, db)
            db_tx = models.Transaction(type=tx_type, amount=amount, description=description, unit_id=unit.id if unit else None)
            db.add(db_tx); db.commit(); db.refresh(db_tx)
            reply_text = f"✅ บันทึก{'รายรับ' if action == 'รับ' else 'รายจ่าย'}สำเร็จ!\n💰 จำนวนเงิน: {amount:,.2f} บาท\n📝 รายละเอียด: {description}\n🏢 ธุรกิจ: {unit.name if unit else 'ทั่วไป'}"
            await get_line_bot_api().reply_message(ReplyMessageRequest(reply_token=event.reply_token, messages=[TextMessage(text=reply_text)]))
            continue

    return {"status": "ok"}

@app.post("/notify/billing-reminder")
async def send_billing_reminder(send_line: bool = False, db: Session = Depends(get_db)):
    unpaid_rooms = db.query(models.DormRoom).filter(models.DormRoom.payment_status != "paid", models.DormRoom.tenant != "", models.DormRoom.tenant != None).all()
    unpaid_houses = db.query(models.RentalHouse).filter(models.RentalHouse.payment_status != "paid", models.RentalHouse.tenant_name != "", models.RentalHouse.tenant_name != None).all()
    unpaid_jobs = db.query(models.GarageJob).filter(models.GarageJob.payment_status != "paid").all()
    
    success_count = 0
    failed_count = 0
    failed_rooms = []
    
    if send_line:
        for room in unpaid_rooms:
            line_user_id = get_line_user_id_for_room(room, db)
            if not line_user_id:
                failed_count += 1
                failed_rooms.append(f"ห้อง {room.number} (ไม่พบ LINE ID)")
                continue
                
            try:
                import json
                current_month_key = get_current_billing_month()
                flex_payload = create_dorm_bill_flex(room, current_month_key)
                
                await get_line_bot_api().push_message(PushMessageRequest(
                    to=line_user_id,
                    messages=[
                        FlexMessage(
                            alt_text=f"ใบแจ้งยอดค่าเช่าห้อง {room.number} ประจำรอบบิล {current_month_key}",
                            contents=FlexContainer.from_json(json.dumps(flex_payload))
                        )
                    ]
                ))
                success_count += 1
            except Exception as e:
                failed_count += 1
                failed_rooms.append(f"ห้อง {room.number} (Error: {str(e)})")
                
    return {
        "status": "success",
        "send_line_executed": send_line,
        "line_push_success": success_count,
        "line_push_failed": failed_count,
        "failed_rooms": failed_rooms,
        "unpaid_rooms": len(unpaid_rooms),
        "unpaid_houses": len(unpaid_houses),
        "unpaid_jobs": len(unpaid_jobs),
        "total_unpaid": len(unpaid_rooms) + len(unpaid_houses) + len(unpaid_jobs)
    }


@app.post("/rooms/apply-daily-fines/")
def apply_daily_fines(force: bool = False, db: Session = Depends(get_db)):
    # Thailand timezone is UTC+7
    from datetime import timezone
    tz_bangkok = timezone(timedelta(hours=7))
    today = datetime.now(tz_bangkok)
    current_day = today.day
    
    # Check if we are inside the fine application period (6th of month to 24th of month)
    # The billing period is 25th to 5th. Fines start accumulating from the 6th onwards.
    if not force and not (6 <= current_day <= 24):
        return {
            "status": "skipped",
            "message": f"ระบบงดเว้นค่าปรับในช่วงรอบจ่ายบิลปกติ (วันที่ 25 ถึง 5 ของเดือนถัดไป) ปัจจุบันคือวันที่ {current_day} (หากต้องการบังคับปรับ ให้ส่งพารามิเตอร์ force=True)",
            "fined_rooms": []
        }
        
    # Get all occupied rooms that have NOT paid
    unpaid_rooms = db.query(models.DormRoom).filter(
        models.DormRoom.payment_status != "paid",
        models.DormRoom.tenant != "",
        models.DormRoom.tenant.isnot(None)
    ).all()
    
    fined_details = []
    for room in unpaid_rooms:
        room.late_days += 1
        room.fine_cost += 100.0
        fined_details.append({
            "room_number": room.number,
            "dorm_key": room.dorm_key,
            "tenant": room.tenant,
            "late_days": room.late_days,
            "fine_cost": room.fine_cost
        })
        
    db.commit()
    
    return {
        "status": "success",
        "message": f"คำนวณและบันทึกค่าปรับสะสมล่าช้าเรียบร้อยแล้ว (บวกเพิ่มวันละ 100 บาท)",
        "fined_count": len(unpaid_rooms),
        "fined_rooms": fined_details
    }

# Maintenance Tickets Endpoints
@app.get("/maintenance-tickets/", response_model=List[schemas.MaintenanceTicket])
def read_maintenance_tickets(db: Session = Depends(get_db)):
    return db.query(models.MaintenanceTicket).order_by(models.MaintenanceTicket.created_at.desc()).all()

@app.patch("/maintenance-tickets/{ticket_id}/", response_model=schemas.MaintenanceTicket)
async def update_maintenance_ticket(ticket_id: int, ticket_update: schemas.MaintenanceTicketUpdate, db: Session = Depends(get_db)):
    ticket = db.query(models.MaintenanceTicket).filter(models.MaintenanceTicket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Maintenance ticket not found")
        
    if ticket_update.status is not None:
        old_status = ticket.status
        ticket.status = ticket_update.status
        if ticket_update.status == "resolved":
            ticket.resolved_at = datetime.utcnow()
        else:
            ticket.resolved_at = None
            
        # Send LINE notification to tenant if status changed and line_user_id is available
        if ticket.line_user_id and old_status != ticket.status:
            status_thai = {
                "pending": "รอดำเนินการ (Pending)",
                "in_progress": "กำลังดำเนินการ (In Progress)",
                "resolved": "แก้ไขเรียบร้อยแล้ว (Resolved)"
            }.get(ticket.status, ticket.status)
            
            status_message = ""
            if ticket.status == "in_progress":
                status_message = (
                    f"🔧 อัปเดตสถานะแจ้งซ่อม ใบงาน #{ticket.id} 🔧\n"
                    f"━━━━━━━━━━━━━━━━━━\n"
                    f"🏢 ห้องพัก: ห้อง {ticket.room_number}\n"
                    f"🔧 รายการชำรุด: {ticket.description}\n"
                    f"⚙️ สถานะใหม่: {status_thai}\n"
                    f"━━━━━━━━━━━━━━━━━━\n"
                    f"ช่างผู้เชี่ยวชาญกำลังเข้าไปดำเนินการตรวจสอบและแก้ไขปัญหาให้ท่านแล้วครับ ขออภัยในความไม่สะดวกครับ 🙏"
                )
            elif ticket.status == "resolved":
                status_message = (
                    f"✅ ปัญหาของคุณได้รับการแก้ไขแล้ว! ใบงาน #{ticket.id} ✅\n"
                    f"━━━━━━━━━━━━━━━━━━\n"
                    f"🏢 ห้องพัก: ห้อง {ticket.room_number}\n"
                    f"🔧 รายการชำรุด: {ticket.description}\n"
                    f"⚙️ สถานะใหม่: {status_thai}\n"
                    f"━━━━━━━━━━━━━━━━━━\n"
                    f"ทีมงานได้ทำการซ่อมแซมและแก้ไขปัญหาดังกล่าวเรียบร้อยแล้ว หากพบข้อบกพร่องหรือมีคำขอเพิ่มเติม สามารถพิมพ์แจ้งเข้ามาได้ทันทีครับ ขอบคุณที่ใช้บริการครับ 😊"
                )
                
            if status_message:
                try:
                    await get_line_bot_api().push_message(PushMessageRequest(
                        to=ticket.line_user_id,
                        messages=[TextMessage(text=status_message)]
                    ))
                except Exception as e:
                    print(f"Failed to send LINE status update to user {ticket.line_user_id}: {e}")
                    
    db.commit()
    db.refresh(ticket)
    return ticket

@app.delete("/maintenance-tickets/{ticket_id}/")
def delete_maintenance_ticket(ticket_id: int, db: Session = Depends(get_db)):
    ticket = db.query(models.MaintenanceTicket).filter(models.MaintenanceTicket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Maintenance ticket not found")
    try:
        db.delete(ticket)
        db.commit()
        return {"status": "success", "message": f"ลบใบงานแจ้งซ่อม #{ticket_id} เรียบร้อยแล้ว"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"เกิดข้อผิดพลาดในการลบใบแจ้งซ่อม: {str(e)}")

# Dormitory Endpoints
@app.get("/rooms/", response_model=List[schemas.DormRoom])
def read_rooms(db: Session = Depends(get_db)):
    return db.query(models.DormRoom).order_by(models.DormRoom.number.asc()).all()

@app.patch("/rooms/{dorm_key}/{number}/", response_model=schemas.DormRoom)
def update_room(dorm_key: str, number: str, room_update: schemas.DormRoomUpdate, db: Session = Depends(get_db)):
    print(f"DEBUG: Received update request for room {number} in {dorm_key}")
    print(f"DEBUG: Payload: {room_update.dict(exclude_unset=True)}")
    room = db.query(models.DormRoom).filter(models.DormRoom.dorm_key == dorm_key, models.DormRoom.number == number).first()
    if not room: raise HTTPException(status_code=404, detail="Room not found")
    old_status = room.payment_status
    try:
        for key, value in room_update.dict(exclude_unset=True).items(): 
            setattr(room, key, value)
        db.commit(); db.refresh(room)
        print(f"DEBUG: Update successful for room {number}")
    except Exception as e:
        db.rollback()
        print(f"DEBUG: Error updating room: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    
    current_month_key = get_current_billing_month(room.payment_date)
    ref_id = f"dorm_payment_{room.id}_{current_month_key}"
    
    # Always keep DormPayment record in sync or create if it doesn't exist
    pmt = db.query(models.DormPayment).filter(models.DormPayment.room_id == room.id, models.DormPayment.month == current_month_key).first()
    if not pmt:
        pmt = models.DormPayment(
            room_id=room.id,
            month=current_month_key,
            payment_status=room.payment_status,
            paid_at=datetime.utcnow() if room.payment_status == "paid" else None
        )
        db.add(pmt)
    
    # Sync all fields from room state to DormPayment snapshot
    pmt.amount = room.rate or 0.0
    pmt.water_cost = room.water_cost or 0.0
    pmt.electric_cost = room.electric_cost or 0.0
    pmt.cleaning_fee = room.cleaning_fee or 0.0
    pmt.other_fee = room.other_fee or 0.0
    pmt.fine_cost = room.fine_cost or 0.0
    pmt.water_meter_prev = room.water_meter_prev or 0.0
    pmt.water_meter = room.water_meter or 0.0
    pmt.electricity_meter_prev = room.electricity_meter_prev or 0.0
    pmt.electricity_meter = room.electricity_meter or 0.0
    pmt.remark = room.remark
    pmt.move_out = room.move_out
    pmt.vacant = room.vacant
    pmt.payment_status = room.payment_status
    if room.payment_status == "paid" and not pmt.paid_at:
        pmt.paid_at = datetime.utcnow()
    elif room.payment_status != "paid":
        pmt.paid_at = None
        
    # Transaction Ledger Management
    total_bill = (room.rate or 0.0) + (room.water_cost or 0.0) + (room.electric_cost or 0.0) + (room.cleaning_fee or 0.0) + (room.other_fee or 0.0) + (room.fine_cost or 0.0)
    
    if room.payment_status == "paid" and old_status != "paid" and room.tenant:
        if total_bill > 0:
            unit = db.query(models.BusinessUnit).filter(models.BusinessUnit.type == models.UnitType.DORMITORY).first()
            if not db.query(models.Transaction).filter(models.Transaction.reference_id == ref_id).first():
                db.add(models.Transaction(type=models.TransactionType.INCOME, amount=total_bill, description=f"ค่าเช่าห้อง {room.number} รอบเดือน {current_month_key} - {room.tenant}", reference_id=ref_id, unit_id=unit.id if unit else None))
    elif room.payment_status != "paid" and old_status == "paid":
        tx = db.query(models.Transaction).filter(models.Transaction.reference_id == ref_id).first()
        if tx:
            db.delete(tx)
            
    db.commit()
    db.refresh(room)
    return room

@app.post("/rooms/rollover/")
def rollover_rooms(db: Session = Depends(get_db)):
    rooms = db.query(models.DormRoom).all()
    for r in rooms:
        current_month_key = get_current_billing_month(r.payment_date)
        
        # Ensure a historical snapshot exists for the month before we reset
        pmt = db.query(models.DormPayment).filter(models.DormPayment.room_id == r.id, models.DormPayment.month == current_month_key).first()
        if not pmt:
            pmt = models.DormPayment(
                room_id=r.id,
                month=current_month_key
            )
            db.add(pmt)
            
        # Freeze current state to snapshot
        pmt.amount = r.rate or 0.0
        pmt.water_cost = r.water_cost or 0.0
        pmt.electric_cost = r.electric_cost or 0.0
        pmt.cleaning_fee = r.cleaning_fee or 0.0
        pmt.other_fee = r.other_fee or 0.0
        pmt.fine_cost = r.fine_cost or 0.0
        pmt.water_meter_prev = r.water_meter_prev or 0.0
        pmt.water_meter = r.water_meter or 0.0
        pmt.electricity_meter_prev = r.electricity_meter_prev or 0.0
        pmt.electricity_meter = r.electricity_meter or 0.0
        pmt.remark = r.remark
        pmt.move_out = r.move_out
        pmt.vacant = r.vacant
        pmt.payment_status = r.payment_status
        if r.payment_status == "paid" and not pmt.paid_at:
            pmt.paid_at = datetime.utcnow()
            
        # Move current meters to previous on room object
        r.water_meter_prev = r.water_meter
        r.electricity_meter_prev = r.electricity_meter
        
        # Reset costs and status for new month
        r.water_cost = 0.0
        r.electric_cost = 0.0
        r.cleaning_fee = 0.0
        r.other_fee = 0.0
        r.late_days = 0
        r.fine_cost = 0.0
        r.payment_status = "pending"
        r.payment_date = ""
        r.remark = ""
        r.move_out = ""
        r.vacant = ""
        
    db.commit()
    return {"status": "success", "message": "ขึ้นรอบบิลใหม่เรียบร้อย!"}

@app.post("/rooms/reset/")
def reset_rooms(db: Session = Depends(get_db)):
    db.close()
    models.Base.metadata.drop_all(bind=engine); models.Base.metadata.create_all(bind=engine)
    seed_business_units(); seed_rooms_and_houses()
    return {"status": "success", "message": "รีเซ็ตข้อมูลทั้งหมดเรียบร้อยแล้ว!"}

# Excel Spreadsheet Endpoints
from pydantic import BaseModel

class SpreadsheetUpdatePayload(BaseModel):
    rate: Optional[float] = None
    water_meter_prev: Optional[float] = None
    water_meter: Optional[float] = None
    electricity_meter_prev: Optional[float] = None
    electricity_meter: Optional[float] = None
    cleaning_fee: Optional[float] = None
    other_fee: Optional[float] = None
    remark: Optional[str] = None
    move_out: Optional[str] = None
    vacant: Optional[str] = None
    payment_status: Optional[str] = None

@app.get("/rooms/spreadsheet/{month}/", response_model=List[schemas.SpreadsheetRoomResponse])
def get_room_spreadsheet(month: str, db: Session = Depends(get_db)):
    active_month = get_current_billing_month()
    rooms = db.query(models.DormRoom).order_by(models.DormRoom.number.asc()).all()
    
    result = []
    for r in rooms:
        if month == active_month:
            pmt = db.query(models.DormPayment).filter(models.DormPayment.room_id == r.id, models.DormPayment.month == month).first()
            result.append({
                "room_id": r.id,
                "number": r.number,
                "floor": r.floor,
                "dorm_key": r.dorm_key,
                "rate": r.rate or 0.0,
                "tenant": r.tenant or "",
                "water_meter_prev": r.water_meter_prev or 0.0,
                "water_meter": r.water_meter or 0.0,
                "electricity_meter_prev": r.electricity_meter_prev or 0.0,
                "electricity_meter": r.electricity_meter or 0.0,
                "water_cost": r.water_cost or 0.0,
                "electric_cost": r.electric_cost or 0.0,
                "cleaning_fee": r.cleaning_fee or 0.0,
                "other_fee": r.other_fee or 0.0,
                "fine_cost": r.fine_cost or 0.0,
                "payment_status": r.payment_status or "pending",
                "remark": r.remark or "",
                "move_out": r.move_out or "",
                "vacant": r.vacant or "",
                "paid_at": pmt.paid_at.isoformat() if pmt and pmt.paid_at else None
            })
        else:
            pmt = db.query(models.DormPayment).filter(models.DormPayment.room_id == r.id, models.DormPayment.month == month).first()
            if pmt:
                result.append({
                    "room_id": r.id,
                    "number": r.number,
                    "floor": r.floor,
                    "dorm_key": r.dorm_key,
                    "rate": pmt.amount or 0.0,
                    "tenant": r.tenant if pmt.payment_status == "paid" else (r.tenant or ""),
                    "water_meter_prev": pmt.water_meter_prev or 0.0,
                    "water_meter": pmt.water_meter or 0.0,
                    "electricity_meter_prev": pmt.electricity_meter_prev or 0.0,
                    "electricity_meter": pmt.electricity_meter or 0.0,
                    "water_cost": pmt.water_cost or 0.0,
                    "electric_cost": pmt.electric_cost or 0.0,
                    "cleaning_fee": pmt.cleaning_fee or 0.0,
                    "other_fee": pmt.other_fee or 0.0,
                    "fine_cost": pmt.fine_cost or 0.0,
                    "payment_status": pmt.payment_status or "unpaid",
                    "remark": pmt.remark or "",
                    "move_out": pmt.move_out or "",
                    "vacant": pmt.vacant or "",
                    "paid_at": pmt.paid_at.isoformat() if pmt.paid_at else None
                })
            else:
                result.append({
                    "room_id": r.id,
                    "number": r.number,
                    "floor": r.floor,
                    "dorm_key": r.dorm_key,
                    "rate": r.rate or 0.0,
                    "tenant": "",
                    "water_meter_prev": 0.0,
                    "water_meter": 0.0,
                    "electricity_meter_prev": 0.0,
                    "electricity_meter": 0.0,
                    "water_cost": 0.0,
                    "electric_cost": 0.0,
                    "cleaning_fee": 0.0,
                    "other_fee": 0.0,
                    "fine_cost": 0.0,
                    "payment_status": "unpaid",
                    "remark": "",
                    "move_out": "",
                    "vacant": "ว่าง",
                    "paid_at": None
                })
    return result

@app.patch("/rooms/spreadsheet/{month}/{room_id}/")
def patch_room_spreadsheet(month: str, room_id: int, payload: SpreadsheetUpdatePayload, db: Session = Depends(get_db)):
    active_month = get_current_billing_month()
    room = db.query(models.DormRoom).filter(models.DormRoom.id == room_id).first()
    if not room: raise HTTPException(status_code=404, detail="Room not found")
    
    water_prev = payload.water_meter_prev if payload.water_meter_prev is not None else room.water_meter_prev
    water_curr = payload.water_meter if payload.water_meter is not None else room.water_meter
    elec_prev = payload.electricity_meter_prev if payload.electricity_meter_prev is not None else room.electricity_meter_prev
    elec_curr = payload.electricity_meter if payload.electricity_meter is not None else room.electricity_meter
    
    water_units = max(0.0, water_curr - water_prev)
    elec_units = max(0.0, elec_curr - elec_prev)
    
    water_cost = water_units * 17.0
    elec_cost = elec_units * 7.0
    
    if month == active_month:
        if payload.rate is not None: room.rate = payload.rate
        if payload.water_meter_prev is not None: room.water_meter_prev = payload.water_meter_prev
        if payload.water_meter is not None: room.water_meter = payload.water_meter
        if payload.electricity_meter_prev is not None: room.electricity_meter_prev = payload.electricity_meter_prev
        if payload.electricity_meter is not None: room.electricity_meter = payload.electricity_meter
        if payload.cleaning_fee is not None: room.cleaning_fee = payload.cleaning_fee
        if payload.other_fee is not None: room.other_fee = payload.other_fee
        if payload.remark is not None: room.remark = payload.remark
        if payload.move_out is not None: room.move_out = payload.move_out
        if payload.vacant is not None: room.vacant = payload.vacant
        
        room.water_cost = water_cost
        room.electric_cost = elec_cost
        
        if payload.payment_status is not None:
            old_status = room.payment_status
            room.payment_status = payload.payment_status
            if payload.payment_status == "paid" and old_status != "paid":
                room.payment_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            elif payload.payment_status != "paid" and old_status == "paid":
                room.payment_date = ""
                
        db.commit()
        db.refresh(room)
        
        pmt = db.query(models.DormPayment).filter(models.DormPayment.room_id == room.id, models.DormPayment.month == month).first()
        if not pmt:
            pmt = models.DormPayment(room_id=room.id, month=month)
            db.add(pmt)
            
        pmt.amount = room.rate or 0.0
        pmt.water_cost = room.water_cost or 0.0
        pmt.electric_cost = room.electric_cost or 0.0
        pmt.cleaning_fee = room.cleaning_fee or 0.0
        pmt.other_fee = room.other_fee or 0.0
        pmt.fine_cost = room.fine_cost or 0.0
        pmt.water_meter_prev = room.water_meter_prev or 0.0
        pmt.water_meter = room.water_meter or 0.0
        pmt.electricity_meter_prev = room.electricity_meter_prev or 0.0
        pmt.electricity_meter = room.electricity_meter or 0.0
        pmt.remark = room.remark
        pmt.move_out = room.move_out
        pmt.vacant = room.vacant
        pmt.payment_status = room.payment_status
        if room.payment_status == "paid" and not pmt.paid_at:
            pmt.paid_at = datetime.utcnow()
        elif room.payment_status != "paid":
            pmt.paid_at = None
            
        db.commit()
    else:
        pmt = db.query(models.DormPayment).filter(models.DormPayment.room_id == room.id, models.DormPayment.month == month).first()
        if not pmt:
            pmt = models.DormPayment(room_id=room.id, month=month)
            db.add(pmt)
            
        if payload.rate is not None: pmt.amount = payload.rate
        if payload.water_meter_prev is not None: pmt.water_meter_prev = payload.water_meter_prev
        if payload.water_meter is not None: pmt.water_meter = payload.water_meter
        if payload.electricity_meter_prev is not None: pmt.electricity_meter_prev = payload.electricity_meter_prev
        if payload.electricity_meter is not None: pmt.electricity_meter = payload.electricity_meter
        if payload.cleaning_fee is not None: pmt.cleaning_fee = payload.cleaning_fee
        if payload.other_fee is not None: pmt.other_fee = payload.other_fee
        if payload.remark is not None: pmt.remark = payload.remark
        if payload.move_out is not None: pmt.move_out = payload.move_out
        if payload.vacant is not None: pmt.vacant = payload.vacant
        
        pmt.water_cost = water_cost
        pmt.electric_cost = elec_cost
        
        if payload.payment_status is not None:
            pmt.payment_status = payload.payment_status
            if payload.payment_status == "paid" and not pmt.paid_at:
                pmt.paid_at = datetime.utcnow()
            elif payload.payment_status != "paid":
                pmt.paid_at = None
                
        db.commit()
        
    return {"status": "success", "message": "บันทึกข้อมูลเรียบร้อย!"}

@app.get("/rooms/utility-history-trends/")
def get_utility_history_trends(db: Session = Depends(get_db)):
    current_date = datetime.now()
    months_list = []
    for i in range(11, -1, -1):
        m = current_date.month - i
        y = current_date.year
        while m <= 0:
            m += 12
            y -= 1
        months_list.append(f"{y:04d}-{m:02d}")
        
    trends = []
    for m in months_list:
        payments = db.query(models.DormPayment).filter(models.DormPayment.month == m).all()
        
        total_water_units = 0.0
        total_water_cost = 0.0
        total_elec_units = 0.0
        total_elec_cost = 0.0
        
        for p in payments:
            water_u = max(0.0, (p.water_meter or 0.0) - (p.water_meter_prev or 0.0))
            elec_u = max(0.0, (p.electricity_meter or 0.0) - (p.electricity_meter_prev or 0.0))
            
            total_water_units += water_u
            total_water_cost += (p.water_cost or 0.0)
            total_elec_units += elec_u
            total_elec_cost += (p.electric_cost or 0.0)
            
        trends.append({
            "month": m,
            "water_units": total_water_units,
            "water_cost": total_water_cost,
            "electricity_units": total_elec_units,
            "electricity_cost": total_elec_cost
        })
        
    return trends

# Garage Endpoints
@app.get("/garage/jobs/", response_model=List[schemas.GarageJob])
def read_garage_jobs(db: Session = Depends(get_db)):
    return db.query(models.GarageJob).order_by(models.GarageJob.created_at.desc()).all()

@app.post("/garage/jobs/", response_model=schemas.GarageJob)
def create_garage_job(job: schemas.GarageJobCreate, db: Session = Depends(get_db)):
    db_job = models.GarageJob(**job.dict())
    db.add(db_job); db.commit(); db.refresh(db_job)
    return db_job

@app.patch("/garage/jobs/{job_id}/", response_model=schemas.GarageJob)
def update_garage_job(job_id: int, job_update: schemas.GarageJobUpdate, db: Session = Depends(get_db)):
    db_job = db.query(models.GarageJob).filter(models.GarageJob.id == job_id).first()
    if not db_job: raise HTTPException(status_code=404, detail="Job not found")
    old_status = db_job.payment_status
    for key, value in job_update.dict(exclude_unset=True).items(): setattr(db_job, key, value)
    if db_job.status in ["finished", "picked_up"] and not db_job.finished_at: db_job.finished_at = datetime.utcnow()
    db.commit(); db.refresh(db_job)
    ref_id = f"garage_payment_{db_job.id}"
    if db_job.payment_status == "paid" and old_status != "paid" and db_job.total_cost > 0:
        unit = db.query(models.BusinessUnit).filter(models.BusinessUnit.type == models.UnitType.GARAGE).first()
        if not db.query(models.Transaction).filter(models.Transaction.reference_id == ref_id).first():
            db.add(models.Transaction(type=models.TransactionType.INCOME, amount=db_job.total_cost, description=f"ค่าซ่อมรถ {db_job.license_plate}", reference_id=ref_id, unit_id=unit.id if unit else None))
            db.commit()
    elif db_job.payment_status != "paid" and old_status == "paid":
        tx = db.query(models.Transaction).filter(models.Transaction.reference_id == ref_id).first()
        if tx:
            db.delete(tx)
            db.commit()
    return db_job

@app.delete("/garage/jobs/{job_id}/")
def delete_garage_job(job_id: int, db: Session = Depends(get_db)):
    db_job = db.query(models.GarageJob).filter(models.GarageJob.id == job_id).first()
    if not db_job: raise HTTPException(status_code=404, detail="Job not found")
    ref_id = f"garage_payment_{db_job.id}"
    tx = db.query(models.Transaction).filter(models.Transaction.reference_id == ref_id).first()
    if tx:
        db.delete(tx)
    db.delete(db_job); db.commit()
    return {"status": "success"}

# Rental House Endpoints
@app.get("/houses/", response_model=List[schemas.RentalHouse])
def read_rental_houses(db: Session = Depends(get_db)):
    return db.query(models.RentalHouse).all()

@app.patch("/houses/{house_id}/", response_model=schemas.RentalHouse)
def update_rental_house(house_id: str, house_update: schemas.RentalHouseUpdate, db: Session = Depends(get_db)):
    db_house = db.query(models.RentalHouse).filter(models.RentalHouse.id == house_id).first()
    if not db_house: raise HTTPException(status_code=404, detail="House not found")
    old_status = db_house.payment_status
    for key, value in house_update.dict(exclude_unset=True).items(): setattr(db_house, key, value)
    db.commit(); db.refresh(db_house)
    current_month_key = get_current_billing_month(db_house.last_payment_date)
    ref_id = f"house_payment_{db_house.id}_{current_month_key}"
    
    if db_house.payment_status == "paid" and old_status != "paid" and db_house.tenant_name:
        total = db_house.monthly_rent + db_house.water_bill + db_house.electric_bill
        if total > 0:
            unit = db.query(models.BusinessUnit).filter(models.BusinessUnit.name == db_house.name).first() or db.query(models.BusinessUnit).filter(models.BusinessUnit.type == models.UnitType.HOUSE).first()
            if not db.query(models.Transaction).filter(models.Transaction.reference_id == ref_id).first():
                db.add(models.Transaction(type=models.TransactionType.INCOME, amount=total, description=f"ค่าเช่าบ้าน {db_house.name} รอบ {current_month_key}", reference_id=ref_id, unit_id=unit.id if unit else None))
            if not db.query(models.HousePayment).filter(models.HousePayment.house_id == db_house.id, models.HousePayment.month == current_month_key).first():
                db.add(models.HousePayment(house_id=db_house.id, month=current_month_key, amount=db_house.monthly_rent, water_bill=db_house.water_bill, electric_bill=db_house.electric_bill, payment_status="paid", paid_at=datetime.utcnow()))
            db.commit()
    elif db_house.payment_status != "paid" and old_status == "paid":
        tx = db.query(models.Transaction).filter(models.Transaction.reference_id == ref_id).first()
        if tx:
            db.delete(tx)
        pmt = db.query(models.HousePayment).filter(models.HousePayment.house_id == db_house.id, models.HousePayment.month == current_month_key).first()
        if pmt:
            db.delete(pmt)
        db.commit()
    return db_house

@app.post("/houses/", response_model=schemas.RentalHouse)
def create_rental_house(house: schemas.RentalHouseCreate, db: Session = Depends(get_db)):
    existing_house = db.query(models.RentalHouse).filter(models.RentalHouse.id == house.id).first()
    if existing_house:
        raise HTTPException(status_code=400, detail=f"มีบ้านเช่ารหัส {house.id} ในระบบแล้ว")
        
    unit = db.query(models.BusinessUnit).filter(models.BusinessUnit.name == house.name).first()
    if not unit:
        unit = models.BusinessUnit(name=house.name, type=models.UnitType.HOUSE)
        db.add(unit)
        db.flush()
        
    db_house = models.RentalHouse(**house.dict())
    try:
        db.add(db_house)
        db.commit()
        db.refresh(db_house)
        return db_house
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"เกิดข้อผิดพลาดในการสร้างบ้านเช่า: {str(e)}")

@app.delete("/houses/{house_id}/")
def delete_rental_house(house_id: str, db: Session = Depends(get_db)):
    db_house = db.query(models.RentalHouse).filter(models.RentalHouse.id == house_id).first()
    if not db_house:
        raise HTTPException(status_code=404, detail="ไม่พบข้อมูลบ้านเช่า")
        
    paid_payments = db.query(models.HousePayment).filter(
        models.HousePayment.house_id == house_id, 
        models.HousePayment.payment_status == "paid"
    ).first()
    
    unit = db.query(models.BusinessUnit).filter(models.BusinessUnit.name == db_house.name).first()
    has_transactions = False
    if unit:
        has_transactions = db.query(models.Transaction).filter(models.Transaction.unit_id == unit.id).first() is not None
        
    if paid_payments or has_transactions:
        raise HTTPException(
            status_code=400, 
            detail="ไม่สามารถลบข้อมูลบ้านเช่านี้ได้ เนื่องจากมีการบันทึกการชำระเงินหรือรายการธุรกรรมทางการเงินในระบบบัญชีแยกประเภทแล้ว เพื่อรักษาความถูกต้องและโปร่งใสของข้อมูลทางบัญชี (Ledger Integrity Safeguard)"
        )
        
    try:
        db.query(models.HousePayment).filter(models.HousePayment.house_id == house_id).delete()
        if unit and not has_transactions:
            db.delete(unit)
        db.delete(db_house)
        db.commit()
        return {"status": "success", "message": f"ลบข้อมูลบ้านเช่า {db_house.name} เรียบร้อยแล้ว"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"เกิดข้อผิดพลาดในการลบข้อมูลบ้านเช่า: {str(e)}")

# Lease Agreement Endpoints
@app.get("/leases/expiring/")
def get_expiring_leases(db: Session = Depends(get_db)):
    rooms = db.query(models.DormRoom).filter(models.DormRoom.lease_end_date != None).all()
    houses = db.query(models.RentalHouse).filter(models.RentalHouse.lease_end_date != None).all()
    expiring = []
    today = datetime.now()
    for r in rooms:
        if not r.lease_end_date: continue
        try:
            # r.lease_end_date is already a datetime object from SQLAlchemy
            end_date = r.lease_end_date
            days = (end_date - today).days
            if days <= 30:
                expiring.append({
                    "id": f"room_{r.id}", "type": "dormitory", "target_name": f"ห้อง {r.number}",
                    "tenant_name": r.tenant or "ไม่มีผู้เช่า",
                    "start_date": r.lease_start_date.strftime("%Y-%m-%d") if r.lease_start_date else None,
                    "end_date": r.lease_end_date.strftime("%Y-%m-%d"),
                    "deposit": r.deposit, "days_left": days,
                    "status": "หมดอายุ" if days < 0 else f"เหลือ {days} วัน"
                })
        except Exception as e:
            print(f"Error processing room lease {r.number}: {e}")
            pass
    for h in houses:
        if not h.lease_end_date: continue
        try:
            end_date = h.lease_end_date
            days = (end_date - today).days
            if days <= 30:
                expiring.append({
                    "id": f"house_{h.id}", "type": "house", "target_name": h.name,
                    "tenant_name": h.tenant_name or "ไม่มีผู้เช่า",
                    "start_date": h.lease_start_date.strftime("%Y-%m-%d") if h.lease_start_date else None,
                    "end_date": h.lease_end_date.strftime("%Y-%m-%d"),
                    "deposit": h.deposit, "days_left": days,
                    "status": "หมดอายุ" if days < 0 else f"เหลือ {days} วัน"
                })
        except Exception as e:
            print(f"Error processing house lease {h.name}: {e}")
            pass
    return expiring

# Payment History
@app.get("/houses/{house_id}/payments/", response_model=List[schemas.HousePayment])
def get_house_payments(house_id: str, db: Session = Depends(get_db)):
    return db.query(models.HousePayment).filter(models.HousePayment.house_id == house_id).order_by(models.HousePayment.month.desc()).all()

@app.get("/dorm-rooms/{room_id}/payments/", response_model=List[schemas.DormPayment])
def get_room_payments(room_id: int, db: Session = Depends(get_db)):
    return db.query(models.DormPayment).filter(models.DormPayment.room_id == room_id).order_by(models.DormPayment.month.desc()).all()

# Business Unit Management
@app.get("/business-units/", response_model=List[schemas.BusinessUnit])
def read_business_units(db: Session = Depends(get_db)): return db.query(models.BusinessUnit).all()

@app.post("/business-units/", response_model=schemas.BusinessUnit)
def create_business_unit(unit: schemas.BusinessUnitCreate, db: Session = Depends(get_db)):
    db_unit = models.BusinessUnit(name=unit.name, type=unit.type); db.add(db_unit); db.commit(); db.refresh(db_unit); return db_unit

@app.put("/business-units/{unit_id}/", response_model=schemas.BusinessUnit)
def update_business_unit(unit_id: int, unit_update: schemas.BusinessUnitCreate, db: Session = Depends(get_db)):
    db_unit = db.query(models.BusinessUnit).filter(models.BusinessUnit.id == unit_id).first()
    if not db_unit: raise HTTPException(status_code=404, detail="Not found")
    db_unit.name = unit_update.name; db_unit.type = unit_update.type; db.commit(); db.refresh(db_unit); return db_unit

@app.delete("/business-units/{unit_id}/")
def delete_business_unit(unit_id: int, db: Session = Depends(get_db)):
    db_unit = db.query(models.BusinessUnit).filter(models.BusinessUnit.id == unit_id).first()
    if not db_unit: raise HTTPException(status_code=404, detail="Not found")
    
    # Cascade check to protect database and bookkeeping integrity
    has_rooms = db.query(models.DormRoom).filter(models.DormRoom.dorm_key == db_unit.name).first()
    has_houses = db.query(models.RentalHouse).filter(models.RentalHouse.name == db_unit.name).first()
    has_invoices = db.query(models.Invoice).filter(models.Invoice.unit_id == unit_id).first()
    has_tx = db.query(models.Transaction).filter(models.Transaction.unit_id == unit_id).first()
    
    if has_rooms or has_houses or has_invoices or has_tx:
        raise HTTPException(status_code=400, detail="ไม่สามารถลบหน่วยธุรกิจนี้ได้ เนื่องจากมีข้อมูลห้องพัก บ้านพัก บิลเรียกเก็บเงิน หรือรายการธุรกรรมเชื่อมโยงอยู่ในระบบบัญชีแยกประเภท (กรุณายกเลิกหรือย้ายข้อมูลที่เกี่ยวข้องก่อนเพื่อรักษาความถูกต้องของข้อมูล)")
        
    db.delete(db_unit); db.commit(); return {"status": "success"}

# Real-time Slip Verification API with SlipOK Integration
@app.post("/payment/verify-slip/")
async def verify_slip(type: str, target_id: str, qr_data: Optional[str] = None, db: Session = Depends(get_db)):
    import random
    import httpx
    
    # 1. Calculate Expected Amount and Get Object References
    expected_amount = 0.0
    current_month_key = get_current_billing_month()
    
    room = None
    house = None
    job = None
    invoice = None
    
    if type == "dorm":
        room = db.query(models.DormRoom).filter(models.DormRoom.id == int(target_id)).first()
        if not room: raise HTTPException(status_code=404, detail="Room not found")
        if room.payment_status == "paid": 
            # Fetch existing ref_no to keep it consistent
            ref_id = f"dorm_payment_{room.id}_{current_month_key}"
            tx = db.query(models.Transaction).filter(models.Transaction.reference_id == ref_id).first()
            return {"status": "already_paid", "amount": room.rate, "ref_no": tx.reference_id if tx else "ALREADY_PAID"}
        expected_amount = room.rate + room.water_cost + room.electric_cost + room.cleaning_fee + room.other_fee + room.fine_cost
        
    elif type == "house":
        house = db.query(models.RentalHouse).filter(models.RentalHouse.id == target_id).first()
        if not house: raise HTTPException(status_code=404, detail="House not found")
        if house.payment_status == "paid":
            ref_id = f"house_payment_{house.id}_{current_month_key}"
            tx = db.query(models.Transaction).filter(models.Transaction.reference_id == ref_id).first()
            return {"status": "already_paid", "amount": house.monthly_rent, "ref_no": tx.reference_id if tx else "ALREADY_PAID"}
        expected_amount = house.monthly_rent + house.water_bill + house.electric_bill
        
    elif type == "garage":
        job = db.query(models.GarageJob).filter(models.GarageJob.id == int(target_id)).first()
        if not job: raise HTTPException(status_code=404, detail="Job not found")
        if job.payment_status == "paid":
            ref_id = f"garage_payment_{job.id}"
            tx = db.query(models.Transaction).filter(models.Transaction.reference_id == ref_id).first()
            return {"status": "already_paid", "amount": job.total_cost, "ref_no": tx.reference_id if tx else "ALREADY_PAID"}
        expected_amount = job.total_cost
        
    elif type == "invoice":
        invoice = db.query(models.Invoice).filter(models.Invoice.id == int(target_id)).first()
        if not invoice: raise HTTPException(status_code=404, detail="Invoice not found")
        if invoice.status == models.InvoiceStatus.PAID:
            ref_id = f"invoice_payment_{invoice.id}"
            tx = db.query(models.Transaction).filter(models.Transaction.reference_id == ref_id).first()
            return {"status": "already_paid", "amount": invoice.amount, "ref_no": tx.reference_id if tx else "ALREADY_PAID"}
        expected_amount = invoice.amount
    else:
        raise HTTPException(status_code=400, detail="Invalid payment type")
        
    # 2. Check SlipOK API Configuration
    slipok_api_key = os.getenv("SLIPOK_API_KEY")
    slipok_branch_id = os.getenv("SLIPOK_BRANCH_ID", "0")
    
    ref_no = ""
    sender_name = "ผู้เช่าในระบบ"
    receiver_name = "หอพัก/บริษัท"
    
    if slipok_api_key:
        # PRODUCTION MODE: Real verification with SlipOK API
        if not qr_data:
            raise HTTPException(
                status_code=400, 
                detail="ระบบอยู่ในโหมดใช้งานจริง (Production Mode) ซึ่งจำเป็นต้องใช้ข้อมูล QR Code จากตัวสลิปจริงเพื่อยืนยันกับ SlipOK API กรุณาสแกนหรือระบุข้อมูล QR Code"
            )
            
        url = f"https://api.slipok.com/api/line/apikey/{slipok_branch_id}"
        headers = {
            "x-authorization": slipok_api_key,
            "Content-Type": "application/json"
        }
        payload = {
            "data": qr_data,
            "amount": float(expected_amount)
        }
        
        try:
            async with httpx.AsyncClient() as client:
                response = await client.post(url, headers=headers, json=payload, timeout=10.0)
            
            if response.status_code != 200:
                raise HTTPException(
                    status_code=400, 
                    detail=f"การเชื่อมต่อกับ SlipOK API ล้มเหลว (HTTP {response.status_code}): {response.text}"
                )
                
            res_json = response.json()
        except httpx.RequestError as exc:
            raise HTTPException(
                status_code=500, 
                detail=f"เกิดข้อผิดพลาดเครือข่ายระหว่างเชื่อมต่อกับ SlipOK API: {str(exc)}"
            )
            
        # Parse Response
        success = res_json.get("success", False)
        data = res_json.get("data", {})
        is_valid = success and (data.get("success", True) if isinstance(data, dict) else True)
        
        if not is_valid:
            error_msg = res_json.get("message") or (data.get("message") if isinstance(data, dict) else None) or "สลิปโอนเงินไม่ผ่านการตรวจสอบความถูกต้อง"
            raise HTTPException(status_code=400, detail=f"การตรวจสอบสลิปล้มเหลว: {error_msg}")
            
        # Extract Verification Details
        ref_no = data.get("transRef")
        if not ref_no:
            raise HTTPException(status_code=400, detail="ไม่พบรหัสอ้างอิงธุรกรรม (Transaction Reference) ในสลิป")
            
        # Recalculate billing month based on actual transaction date if available
        trans_date_str = data.get("transDate")
        if trans_date_str:
            current_month_key = get_current_billing_month(trans_date_str)
            
        # Duplicate Prevention Checklist
        existing_tx = db.query(models.Transaction).filter(models.Transaction.reference_id == ref_no).first()
        if existing_tx:
            raise HTTPException(
                status_code=400, 
                detail="ตรวจพบความเสี่ยงด้านความปลอดภัย: สลิปการโอนเงินนี้ (Ref No.) เคยถูกนำมาบันทึกชำระเงินในระบบไปแล้ว ไม่สามารถใช้งานซ้ำได้เพื่อป้องกันการโกงยอดเงิน"
            )
            
        sender_name = data.get("sender", {}).get("displayName") or data.get("sender", {}).get("name") or "ผู้โอนเงินจริง"
        receiver_name = data.get("receiver", {}).get("displayName") or data.get("receiver", {}).get("name") or "หอพัก/บริษัท"
        amount_paid = data.get("amount", 0.0)
        
        # Verify Amount Correctness
        if abs(float(amount_paid) - float(expected_amount)) > 0.01:
            raise HTTPException(
                status_code=400, 
                detail=f"ยอดเงินในสลิปจริง ({amount_paid:,.2f} บาท) ไม่สอดคล้องกับยอดเงินที่ต้องชำระในระบบ ({expected_amount:,.2f} บาท)"
            )
            
    else:
        # SANDBOX / FALLBACK MODE: Intelligent Simulation with constraints
        ref_no = f"SIM-SLIP-{datetime.now().strftime('%Y%m%d')}-" + "".join([str(random.randint(0, 9)) for _ in range(6)])
        
        while db.query(models.Transaction).filter(models.Transaction.reference_id == ref_no).first():
            ref_no = f"SIM-SLIP-{datetime.now().strftime('%Y%m%d')}-" + "".join([str(random.randint(0, 9)) for _ in range(6)])
            
        sender_name = "ผู้เช่าจำลอง (โหมดทดสอบ)"
        receiver_name = "บริษัท ซอฟเวอเรน โวลต์ จำกัด (โหมดทดสอบ)"
        
    # 3. Apply Bookkeeping Changes & Persist Records
    if type == "dorm" and room:
        room.payment_status = "paid"
        room.payment_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Create consolidated ledger entry using verified ref_no as reference_id
        db.add(models.Transaction(
            type=models.TransactionType.INCOME, 
            amount=expected_amount, 
            description=f"ค่าเช่าห้อง {room.number} รอบ {current_month_key} (ผู้โอน: {sender_name}, อ้างอิงสลิป: {ref_no})", 
            reference_id=ref_no, 
            unit_id=(lambda u: u.id if u else None)(db.query(models.BusinessUnit).filter(models.BusinessUnit.type == models.UnitType.DORMITORY).first())
        ))
            
        if not db.query(models.DormPayment).filter(models.DormPayment.room_id == room.id, models.DormPayment.month == current_month_key).first():
            db.add(models.DormPayment(
                room_id=room.id, 
                month=current_month_key, 
                amount=room.rate, 
                water_cost=room.water_cost, 
                electric_cost=room.electric_cost, 
                cleaning_fee=room.cleaning_fee, 
                other_fee=room.other_fee, 
                fine_cost=room.fine_cost, 
                payment_status="paid", 
                paid_at=datetime.utcnow()
            ))
            
    elif type == "house" and house:
        house.payment_status = "paid"
        house.last_payment_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Create consolidated ledger entry
        db.add(models.Transaction(
            type=models.TransactionType.INCOME, 
            amount=expected_amount, 
            description=f"ค่าเช่าบ้าน {house.name} รอบ {current_month_key} (ผู้โอน: {sender_name}, อ้างอิงสลิป: {ref_no})", 
            reference_id=ref_no, 
            unit_id=(lambda u: u.id if u else None)(db.query(models.BusinessUnit).filter(models.BusinessUnit.name == house.name).first() or db.query(models.BusinessUnit).filter(models.BusinessUnit.type == models.UnitType.HOUSE).first())
        ))
        
        if not db.query(models.HousePayment).filter(models.HousePayment.house_id == house.id, models.HousePayment.month == current_month_key).first():
            db.add(models.HousePayment(
                house_id=house.id, 
                month=current_month_key, 
                amount=house.monthly_rent, 
                water_bill=house.water_bill, 
                electric_bill=house.electric_bill, 
                payment_status="paid", 
                paid_at=datetime.utcnow()
            ))
            
    elif type == "garage" and job:
        job.payment_status = "paid"
        job.status = "picked_up"
        job.finished_at = datetime.utcnow()
        
        # Create consolidated ledger entry
        db.add(models.Transaction(
            type=models.TransactionType.INCOME, 
            amount=expected_amount, 
            description=f"ค่าซ่อมรถ {job.license_plate} (ผู้โอน: {sender_name}, อ้างอิงสลิป: {ref_no})", 
            reference_id=ref_no, 
            unit_id=(lambda u: u.id if u else None)(db.query(models.BusinessUnit).filter(models.BusinessUnit.type == models.UnitType.GARAGE).first())
        ))
        
    elif type == "invoice" and invoice:
        invoice.status = models.InvoiceStatus.PAID
        
        # Create consolidated ledger entry
        db.add(models.Transaction(
            type=models.TransactionType.INCOME, 
            amount=expected_amount, 
            description=f"ชำระบิล: {invoice.title} (ผู้โอน: {sender_name}, อ้างอิงสลิป: {ref_no})", 
            reference_id=ref_no, 
            unit_id=invoice.unit_id, 
            customer_id=invoice.customer_id
        ))
        
    db.commit()
    
    # Send real-time LINE notification to owner
    biz_type_thai = "รายรับการชำระเงิน"
    details_str = ""
    if type == "dorm" and room:
        biz_type_thai = "ค่าเช่าห้องพัก (Dormitory)"
        details_str = f"ห้องพักหมายเลข: {room.number} (ชั้น {room.floor})"
    elif type == "house" and house:
        biz_type_thai = "ค่าเช่าบ้านพัก (Rental House)"
        details_str = f"บ้านเช่า: {house.name}"
    elif type == "garage" and job:
        biz_type_thai = "ค่าบริการซ่อมรถอู่ (Garage Job)"
        details_str = f"รถยนต์: {job.license_plate} ({job.car_model})"
    elif type == "invoice" and invoice:
        biz_type_thai = "บิลทั่วไป (Invoice)"
        details_str = f"รายการ: {invoice.title}"

    time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    msg = (
        f"\n💰 [เงินเข้าผ่าน SlipOK] ตรวจพบยอดชำระสำเร็จ!\n"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"🔹 ธุรกิจ: {biz_type_thai}\n"
        f"🔹 รายละเอียด: {details_str}\n"
        f"🔹 ยอดโอนจริง: {expected_amount:,.2f} บาท\n"
        f"🔹 ผู้โอนเงิน: {sender_name}\n"
        f"🔹 รหัสอ้างอิงสลิป: {ref_no}\n"
        f"🔹 เวลาที่บันทึก: {time_str}"
    )
    await send_financial_alert_to_owner(msg)
    
    return {
        "status": "success", 
        "amount": expected_amount, 
        "ref_no": ref_no,
        "sender": sender_name,
        "receiver": receiver_name,
        "message": "สแกนและตรวจสอบข้อมูลกับระบบธนาคารเรียบร้อยแล้ว" if slipok_api_key else "ตรวจสอบสลิปสำเร็จ (โหมดทดสอบจำลอง Sandbox)"
    }
